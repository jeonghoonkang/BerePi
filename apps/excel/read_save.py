# -*- coding: utf-8 -*-
# Author : jeonghoonkang , https://github.com/jeonghoonkang
# Author : jeongmoon417 , https://github.com/jeongmoon417

# 참고 url
# https://code.tutsplus.com/ko/tutorials/how-to-work-with-excel-documents-using-python--cms-25698
# http://egloos.zum.com/mcchae/v/11120944

import datetime
from openpyxl import Workbook
import openpyxl
import time
# import XlsxWriter
import sys
import importlib
# sys.path.insert(0, '../doc_design')
# (to do) have to find how to add different location directory path and file
# now just using same dir location file

# from openpyxl.workbook import Workbook
# from openpyxl.writer.excel import ExcelWriter
# (error) from openpyxl.cell import get_column_letter
# from openpyxl import load_workbook

global EE_file
global EA_file

class excell_class :
    __ofile = None

    def __init__(self):
        pass

    #@staticmethod
    def open_exc_doc(self,__file):
        # using unicode file name with u syntax
        __ofile = openpyxl.load_workbook(__file)

        #pout = "ix: %s, mdsid: %s, donecheck: %d \r" %(ix, mds_id, donecheck)
        pout = "   ... file opened \n"
        sys.stdout.write(pout)
        sys.stdout.flush()
        return __ofile

    def read_vertical(self, sheet, __start, __end):
        __vertical = []
        #print " ... Please use column[n]:column[m], vertical read "
        cell_of_col = sheet[__start:__end]
        for row in cell_of_col:
            for cell in row:
                v = cell.value
                if v == None: continue # do nothing below code, back to next for loop step
                __vertical.append(v) # 리스트 __vertical에 아이디 추가
        return __vertical #__cnt, __cnt_n # 세로 셀 데이터, 데이터 갯수, None 갯수

    def save_vdata(self,  __list, __lname = None, __ucode = False, _sn = False):
        #리스트 내부, 숫자를 모두 유니코드로 변환
        if __ucode == True: __list = [unicode(i) for i in __list]
        if _sn == True:
            #사업장번호 포맷 201602260257-01 (15문자, -00 형식으로 맞춤)
            _n_list=[]
            for k in __list:
                if k.find('-') == -1 : k = k + '-01'
                elif k.find('-0') == -1 : k = k[:-1] + '0' + k[-1:]
                _n_list.append(k)
            __list = _n_list
        __t = str(datetime.datetime.now())
        __t = ''.join(__t.split())
        __t = __t[:-15]
        __odata = 'save_list='
        if __lname != None : __odata = __lname +'_list='
        __odata = __odata + str(__list)
        filename = '__out_'+__t+'.py'
        if __lname != None : filename = __lname + filename

        __ofile = open(filename,"w")
        print "  file saving ", filename
        __ofile.write(__odata)
        __ofile.close()
        return __list

def find_mdsid(__l1, __id, vnum):

    # 엑셀 읽어서, 동일 아이디 찾아서 index 위치 리턴
    # D:사업장번호 B:미터기번호
    for k in range(0,len(vnum)):
        _tmp = vnum[k]
        if _tmp == None: _tmp = '0'
        if _tmp[:2] != '20': # 사업장번호 2016으로 시작하도록 확인, 변경
            vnum[k] = '20'+_tmp
        #if len(vnum[k]) != 12 : print "  Length is strage ... %d " %len(vnum[k])

    found_id = []
    for k in __l1:
        print
        try:
            print k,
            f_ix = vnum.index(k) # 리스트에 k 와 동일한 숫자 없으면, except 발생
            print "found at (%d)" %f_ix
            found_id.append(__id[f_ix])
        except:
            found_id.append(None)
        time.sleep(0.2)

    return found_id

def merge_list(__tl, __list1, __list2):
    if len(__list1) != len(__list2):
        exit ("[error] Please check the lines of excel file ")
    __tl = [ [__list1[k], __list2[k]] for k in range(len(__list1)) ]
    return __tl

if __name__ == "__main__":

    print "  Working... please input file name, like *anyname.py* "
    volubility = len(sys.argv)
    _t_string_ = " ".join(sys.argv)

    if volubility > 1:
        _t_string = sys.argv[1]
        print "  input file is", _t_string
    else : exit (" We need excel file which contains MDS_ID")

    #exit("  It is OK, debug exit")
    # 입력파일은 실행할때, arg 로 입력함
    E1_file = _t_string

    exl_class = excell_class()
    e1 = exl_class.open_exc_doc(E1_file)
    sheets = e1.get_sheet_names()
    sh1 = e1.get_sheet_by_name(sheets[0])
    lst1 = exl_class.read_vertical(sh1,'a1','a500')
    u1 = exl_class.save_vdata(lst1,'_id_list_',True)

    idlist = u1

    print " Loss ", float(100* idlist.count(None) / len(idlist)), '%', idlist.count(None), len(idlist)
