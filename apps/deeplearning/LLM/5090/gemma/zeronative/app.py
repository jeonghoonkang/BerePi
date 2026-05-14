from __future__ import annotations

import base64
from datetime import datetime
import html
import io
import json
import os
from pathlib import Path
import re
import secrets
import shutil
import subprocess
import time
from typing import Iterable
from urllib.parse import quote, unquote, urljoin, urlparse
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import pandas as pd
from pypdf import PdfReader
import requests
import streamlit as st
import streamlit.components.v1 as components
from PIL import Image

DEFAULT_OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://127.0.0.1:11434")
DEFAULT_MODEL = os.getenv("OLLAMA_MODEL", "gemma3:4b")
REQUEST_TIMEOUT = int(os.getenv("OLLAMA_TIMEOUT", "600"))
MAX_PREVIEW_ROWS = 20
APP_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = APP_DIR / "workspace"
APP_SETTINGS_PATH = APP_DIR / "app_settings.json"
MAX_TOOL_FILE_BYTES = 1_000_000
MAX_TOOL_ROUNDS = 8
MAX_IDENTICAL_TOOL_ROUNDS = 2
MAX_WORKSPACE_SCAN_FILES = 8
MAX_WORKSPACE_SCAN_CHARS = 3_000
MAX_RAG_FILES = 60
MAX_RAG_CHUNKS = 300
MAX_RAG_CHUNK_CHARS = 1400
RAG_TOP_K = 6
WEBDAV_TIMEOUT = 60
WEBDAV_NS = {
    "d": "DAV:",
}

DEFAULT_WEBDAV_READ_PATH = "/remote.php/dav/files/username/"
DEFAULT_WEBDAV_READ_PATH_PLACEHOLDER = "sub dir name"
DEFAULT_WEBDAV_SUBDIR_PLACEHOLDER = "sub dir name"

SUPPORTED_MODEL_OPTIONS = [
    "gemma3:1b",
    "gemma3:4b",
    "gemma3:12b",
    "gemma3:27b",
    "qwen3:32b",
    "qwen3.5:9b",
    "qwen2.5-coder:7b",
    "qwen3-coder:30b",
]

MODEL_MEMORY_GUIDE_GB = {
    "gemma3:1b": 4,
    "gemma3:4b": 8,
    "gemma3:12b": 20,
    "gemma3:27b": 40,
    "qwen3:32b": 20,
    "qwen3.5:9b": 11,
    "qwen2.5-coder:7b": 8,
    "qwen3-coder:30b": 20,
}

MODEL_DEFAULT_TEMPERATURES = {
    "gemma3:1b": 0.7,
    "gemma3:4b": 0.7,
    "gemma3:12b": 0.6,
    "gemma3:27b": 0.5,
    "qwen3:32b": 0.4,
    "qwen3.5:9b": 0.3,
    "qwen2.5-coder:7b": 0.2,
    "qwen3-coder:30b": 0.2,
}

FILE_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": "List files and directories inside the workspace directory.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Relative path inside the workspace", "default": "."},
                    "recursive": {"type": "boolean", "description": "Whether to include nested files", "default": False},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "Read a UTF-8 text file from the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["path"],
                "properties": {
                    "path": {"type": "string", "description": "Relative file path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": "Write UTF-8 text content to a file inside the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["path", "content"],
                "properties": {
                    "path": {"type": "string", "description": "Relative file path inside the workspace"},
                    "content": {"type": "string", "description": "Text content to write to the file"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "copy_path",
            "description": "Copy a file or directory inside the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["source_path", "destination_path"],
                "properties": {
                    "source_path": {"type": "string", "description": "Relative source path inside the workspace"},
                    "destination_path": {"type": "string", "description": "Relative destination path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_path",
            "description": "Delete a file or directory inside the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["path"],
                "properties": {
                    "path": {"type": "string", "description": "Relative path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "download_file",
            "description": "Prepare a file inside the workspace directory for user download in the UI.",
            "parameters": {
                "type": "object",
                "required": ["path"],
                "properties": {
                    "path": {"type": "string", "description": "Relative file path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_workbook_info",
            "description": "Get workbook sheet names and basic dimensions for an Excel file in the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_sheet_preview",
            "description": "Preview the first rows of a sheet in an Excel file from the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to preview"},
                    "max_rows": {"type": "integer", "description": "Number of rows to preview", "default": 20},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_read_cells",
            "description": "Read a cell or cell range from an Excel sheet in the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name", "cell_range"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to read"},
                    "cell_range": {"type": "string", "description": "Excel range like A1 or A1:C10"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_write_cell",
            "description": "Write a value into a single Excel cell in the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name", "cell", "value"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to update"},
                    "cell": {"type": "string", "description": "Target cell like B3"},
                    "value": {"type": "string", "description": "Value to write"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_aggregate_range",
            "description": "Run numeric aggregation on an Excel range. Supported operations: sum, average, min, max, count.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name", "cell_range", "operation"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to aggregate"},
                    "cell_range": {"type": "string", "description": "Excel range like A1:C10"},
                    "operation": {"type": "string", "description": "sum, average, min, max, or count"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_calculate_statistics",
            "description": "Read an Excel sheet into a dataframe, optionally filter rows, and calculate statistics for one or more columns.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to analyze"},
                    "target_columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Columns to analyze. If omitted, numeric columns are used.",
                    },
                    "group_by": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Optional columns used for grouped statistics.",
                    },
                    "filters": {
                        "type": "array",
                        "description": "Optional row filters such as column/operator/value rules.",
                        "items": {
                            "type": "object",
                            "required": ["column", "operator", "value"],
                            "properties": {
                                "column": {"type": "string"},
                                "operator": {"type": "string", "description": "One of eq, ne, gt, gte, lt, lte, contains"},
                                "value": {},
                            },
                        },
                    },
                    "statistics": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Statistics to calculate. Supported: count, sum, mean, min, max, median, std, nunique.",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_merge_files",
            "description": "Merge multiple Excel files from the workspace into one Excel file. Supported modes: append_rows, separate_sheets.",
            "parameters": {
                "type": "object",
                "required": ["source_paths", "output_path"],
                "properties": {
                    "source_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Relative Excel file paths inside the workspace",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Relative output Excel file path inside the workspace",
                    },
                    "mode": {
                        "type": "string",
                        "description": "append_rows or separate_sheets",
                        "default": "append_rows",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_stack_files_to_single_sheet",
            "description": "Read multiple Excel files and stack their sheet data into one worksheet vertically with blank row gaps between blocks.",
            "parameters": {
                "type": "object",
                "required": ["source_paths", "output_path"],
                "properties": {
                    "source_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Relative Excel file paths inside the workspace",
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Relative output Excel file path inside the workspace",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Output worksheet name",
                        "default": "MergedData",
                    },
                    "gap_rows": {
                        "type": "integer",
                        "description": "Number of blank rows between file blocks",
                        "default": 2,
                    },
                },
            },
        },
    },
]


def load_app_settings() -> dict:
    """Load persisted app settings from the local settings file."""
    if not APP_SETTINGS_PATH.exists():
        return {}
    try:
        data = json.loads(APP_SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def save_app_settings(settings: dict) -> None:
    """Persist app settings to the local settings file."""
    APP_SETTINGS_PATH.write_text(
        json.dumps(settings, ensure_ascii=True, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def get_access_control_settings() -> dict:
    """Return server-side login settings for the app access gate."""
    settings = load_app_settings()
    access_control = settings.get("access_control") if isinstance(settings.get("access_control"), dict) else {}

    login_id = os.getenv("ZERONATIVE_APP_LOGIN_ID")
    if not login_id:
        configured_id = access_control.get("login_id")
        if isinstance(configured_id, str) and configured_id.strip():
            login_id = configured_id.strip()

    login_password = os.getenv("ZERONATIVE_APP_LOGIN_PASSWORD")
    if not login_password:
        configured_password = access_control.get("login_password")
        if isinstance(configured_password, str) and configured_password.strip():
            login_password = configured_password

    return {
        "login_id": login_id.strip() if isinstance(login_id, str) else "",
        "login_password": login_password if isinstance(login_password, str) else "",
        "is_configured": bool(isinstance(login_id, str) and login_id.strip() and isinstance(login_password, str) and login_password),
    }


def render_access_gate() -> bool:
    """Require login before rendering the main application."""
    if "app_authenticated" not in st.session_state:
        st.session_state.app_authenticated = False
    if "app_auth_error" not in st.session_state:
        st.session_state.app_auth_error = ""

    if st.session_state.app_authenticated:
        return True

    credentials = get_access_control_settings()

    login_col, spacer_col = st.columns([1.2, 1.8], gap="large")
    with login_col:
        st.title("ZeroNative AI for RTX 5090")
        if not credentials["is_configured"]:
            st.error("Access control is not configured. Set login credentials on the server before using this app.")
        else:
            st.caption("Sign in to access this page.")
            with st.form("app_access_login", clear_on_submit=False):
                login_id = st.text_input("ID", key="app_access_login_id")
                login_password = st.text_input("Password", type="password", key="app_access_login_password")
                submitted = st.form_submit_button("Sign In", use_container_width=True)

            if submitted:
                login_ok = secrets.compare_digest(login_id.strip(), credentials["login_id"])
                password_ok = secrets.compare_digest(login_password, credentials["login_password"])
                if login_ok and password_ok:
                    st.session_state.app_authenticated = True
                    st.session_state.app_auth_error = ""
                    st.rerun()
                else:
                    st.session_state.app_auth_error = "Invalid ID or password."

            if st.session_state.app_auth_error:
                st.error(st.session_state.app_auth_error)

    with spacer_col:
        if not credentials["is_configured"]:
            st.markdown(
                """
                ### Access Control Required
                This page will not start with a default login anymore.

                Configure credentials on the server with one of these methods:
                - Environment variables: `ZERONATIVE_APP_LOGIN_ID`, `ZERONATIVE_APP_LOGIN_PASSWORD`
                - `app_settings.json` entry: `access_control.login_id`, `access_control.login_password`

                After credentials are configured, refresh the page and sign in.
                """
            )
        else:
            st.markdown(
                """
                ### Access Control
                This page now requires a login before the application UI is shown.

                Configure credentials on the server with one of these methods:
                - Environment variables: `ZERONATIVE_APP_LOGIN_ID`, `ZERONATIVE_APP_LOGIN_PASSWORD`
                - `app_settings.json` entry: `access_control.login_id`, `access_control.login_password`
                """
            )

    return False


def get_saved_model_root() -> str | None:
    """Return the saved model storage root path, if present."""
    value = load_app_settings().get("ollama_models_path")
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def persist_model_root(path_value: str) -> str:
    """Store the chosen model storage root and return its normalized value."""
    normalized = str(Path(path_value).expanduser())
    settings = load_app_settings()
    settings["ollama_models_path"] = normalized
    save_app_settings(settings)
    return normalized


def get_saved_webdav_settings() -> dict:
    """Return persisted WebDAV settings with sane defaults."""
    settings = load_app_settings()
    webdav_settings = settings.get("webdav") if isinstance(settings.get("webdav"), dict) else {}
    read_paths = webdav_settings.get(
        "read_paths",
        [""] * 4,
    )
    if not isinstance(read_paths, list):
        read_paths = [""] * 4
    normalized_paths = [str(value).strip() for value in read_paths[:4]]
    while len(normalized_paths) < 4:
        normalized_paths.append("")
    return {
        "base_url": str(webdav_settings.get("base_url", "")).strip(),
        "username": str(webdav_settings.get("username", "")).strip(),
        "password": "",
        "read_paths": normalized_paths,
        "subdir_path": str(webdav_settings.get("subdir_path", "")).strip(),
    }


def persist_webdav_settings(base_url: str, username: str, read_paths: list[str], subdir_path: str) -> None:
    """Persist non-secret WebDAV connection settings to the local settings file."""
    settings = load_app_settings()
    settings["webdav"] = {
        "base_url": base_url.strip(),
        "username": username.strip(),
        "read_paths": [path.strip() for path in read_paths[:4]],
        "subdir_path": subdir_path.strip(),
    }
    save_app_settings(settings)


def render_client_webdav_storage_helper() -> None:
    """Persist WebDAV form values in the browser localStorage instead of on the server."""
    components.html(
        """
        <script>
        (function() {
          const storageKey = "berepi.zeronative.webdav_settings";
          const fieldNames = [
            "WebDAV Base URL",
            "WebDAV Username",
            "WebDAV Password / App Token",
            "Read Path 1",
            "Read Path 2",
            "Read Path 3",
            "Read Path 4",
            "Subdir Path",
          ];

          function loadSavedValues() {
            try {
              return JSON.parse(window.localStorage.getItem(storageKey) || "{}");
            } catch (error) {
              return {};
            }
          }

          function saveValues(values) {
            try {
              window.localStorage.setItem(storageKey, JSON.stringify(values));
            } catch (error) {
              console.warn("Failed to save WebDAV settings in localStorage", error);
            }
          }

          function findInput(label) {
            return window.parent.document.querySelector(`input[aria-label="${label}"]`);
          }

          function dispatchInputEvents(element) {
            element.dispatchEvent(new Event("input", { bubbles: true }));
            element.dispatchEvent(new Event("change", { bubbles: true }));
            element.dispatchEvent(new KeyboardEvent("keydown", { bubbles: true, key: "Enter", code: "Enter" }));
          }

          function applySavedValues() {
            const saved = loadSavedValues();
            fieldNames.forEach((label) => {
              const element = findInput(label);
              if (!element) {
                return;
              }
              if (saved[label] && element.value !== saved[label]) {
                element.value = saved[label];
                dispatchInputEvents(element);
              }
            });
          }

          function bindPersistence() {
            const saved = loadSavedValues();
            fieldNames.forEach((label) => {
              const element = findInput(label);
              if (!element || element.dataset.webdavBound === "true") {
                return;
              }
              element.dataset.webdavBound = "true";
              element.addEventListener("input", () => {
                saved[label] = element.value || "";
                saveValues(saved);
              });
              element.addEventListener("change", () => {
                saved[label] = element.value || "";
                saveValues(saved);
              });
            });
          }

          setTimeout(() => {
            applySavedValues();
            bindPersistence();
          }, 250);

          const observer = new MutationObserver(() => {
            bindPersistence();
          });
          observer.observe(window.parent.document.body, { childList: true, subtree: true });
        })();
        </script>
        """,
        height=0,
        width=0,
    )


def excel_to_context(uploaded_file) -> str:
    """Create a concise, prompt-friendly summary from an Excel workbook."""
    uploaded_file.seek(0)
    workbook = pd.ExcelFile(uploaded_file)
    sections: list[str] = []

    for sheet_name in workbook.sheet_names:
        frame = workbook.parse(sheet_name)
        preview = frame.head(MAX_PREVIEW_ROWS).fillna("")
        preview_csv = preview.to_csv(index=False)
        sections.append(
            "\n".join(
                [
                    f"[Sheet] {sheet_name}",
                    f"Rows: {len(frame)}",
                    f"Columns: {len(frame.columns)}",
                    "Column names: " + ", ".join(str(column) for column in frame.columns),
                    f"Preview ({min(len(preview), MAX_PREVIEW_ROWS)} rows):",
                    preview_csv.strip(),
                ]
            )
        )

    return "\n\n".join(sections)


def tokenize_query(text: str) -> list[str]:
    """Tokenize user text for simple lexical RAG ranking."""
    return [token for token in re.findall(r"[0-9A-Za-z_가-힣]{2,}", text.lower())]


def normalize_webdav_base_url(base_url: str) -> str:
    """Normalize a WebDAV origin URL like https://host:port."""
    cleaned = base_url.strip()
    if not cleaned:
        return ""
    parsed = urlparse(cleaned if "://" in cleaned else f"https://{cleaned}")
    if not parsed.scheme or not parsed.netloc:
        return cleaned.rstrip("/")
    return f"{parsed.scheme}://{parsed.netloc}"


def normalize_webdav_read_path(base_url: str, read_path: str) -> str:
    """Normalize a user-provided WebDAV root path into an absolute server path."""
    cleaned_input = read_path.strip()
    if not cleaned_input:
        return ""

    if cleaned_input.startswith("http://") or cleaned_input.startswith("https://"):
        parsed_input = urlparse(cleaned_input)
        return unquote(parsed_input.path).rstrip("/") or "/"

    decoded_input = unquote(cleaned_input).strip()
    if not decoded_input.startswith("/"):
        decoded_input = f"/{decoded_input}"
    return decoded_input.rstrip("/") or "/"


def normalize_webdav_subdir_path(subdir_path: str) -> str:
    """Normalize an optional subdirectory path relative to a read root."""
    return unquote(subdir_path).strip().strip("/")


def join_webdav_paths(root_path: str, subdir_path: str) -> str:
    """Join a WebDAV root path and an optional subdirectory path."""
    normalized_root = normalize_webdav_read_path("", root_path)
    normalized_subdir = normalize_webdav_subdir_path(subdir_path)
    if not normalized_subdir:
        return normalized_root
    return f"{normalized_root.rstrip('/')}/{normalized_subdir}"


def build_webdav_url(base_url: str, server_path: str) -> str:
    """Build a full WebDAV URL from an origin and absolute server path."""
    normalized_base_url = normalize_webdav_base_url(base_url)
    normalized_server_path = normalize_webdav_read_path("", server_path)
    encoded_path = "/".join(quote(part) for part in normalized_server_path.split("/") if part)
    return f"{normalized_base_url}/{encoded_path}"


def test_webdav_connection(
    base_url: str,
    username: str,
    password: str,
    read_paths: list[str],
    subdir_path: str,
) -> tuple[str, list[str]]:
    """Validate the current WebDAV credentials and one or more configured paths."""
    session = requests.Session()
    session.auth = (username, password)

    normalized_base_url = normalize_webdav_base_url(base_url)
    normalized_paths = [join_webdav_paths(path, subdir_path) for path in read_paths if path.strip()]
    if not normalized_paths:
        normalized_paths = [""]

    status_lines: list[str] = []
    for normalized_path in normalized_paths:
        target_url = build_webdav_url(normalized_base_url, normalized_path)
        try:
            listing_root = webdav_propfind(session, target_url, depth="0")
            entries = parse_webdav_listing(listing_root, normalized_path)
            status_lines.append(f"OK: {target_url} ({len(entries)} item metadata row(s) returned)")
        except requests.HTTPError as exc:
            detail = exc.response.text.strip()[:300] if exc.response is not None and exc.response.text else str(exc)
            raise RuntimeError(f"Connection test failed for {target_url}: {detail}") from exc

    return "WebDAV connection test passed.", status_lines


def webdav_propfind(session: requests.Session, target_url: str, depth: str = "1") -> ET.Element:
    """Run a WebDAV PROPFIND request and return the XML root."""
    response = session.request(
        "PROPFIND",
        target_url,
        headers={"Depth": depth},
        data=(
            '<?xml version="1.0" encoding="utf-8" ?>'
            '<d:propfind xmlns:d="DAV:"><d:prop><d:resourcetype/><d:getcontentlength/></d:prop></d:propfind>'
        ),
        timeout=WEBDAV_TIMEOUT,
    )
    response.raise_for_status()
    return ET.fromstring(response.text)


def parse_webdav_listing(xml_root: ET.Element, base_path: str) -> list[dict]:
    """Parse a PROPFIND XML response into file and directory entries."""
    entries: list[dict] = []
    normalized_base_path = normalize_webdav_read_path("", base_path)

    for response_element in xml_root.findall("d:response", WEBDAV_NS):
        href = response_element.findtext("d:href", default="", namespaces=WEBDAV_NS)
        if not href:
            continue

        decoded_href = unquote(href)
        href_path = urlparse(decoded_href).path
        relative_path = href_path
        if normalized_base_path and href_path.startswith(normalized_base_path):
            relative_path = href_path[len(normalized_base_path):]
        relative_path = relative_path.strip("/")

        prop = response_element.find("d:propstat/d:prop", WEBDAV_NS)
        if prop is None:
            continue

        is_collection = prop.find("d:resourcetype/d:collection", WEBDAV_NS) is not None
        size_text = prop.findtext("d:getcontentlength", default="", namespaces=WEBDAV_NS).strip()
        try:
            size = int(size_text) if size_text else None
        except ValueError:
            size = None

        entries.append(
            {
                "href": decoded_href,
                "full_path": href_path,
                "relative_path": relative_path,
                "is_collection": is_collection,
                "size": size,
            }
        )

    return entries


def extract_pdf_text(content: bytes) -> str:
    """Extract text from a PDF byte payload."""
    reader = PdfReader(io.BytesIO(content))
    pages: list[str] = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n".join(page for page in pages if page.strip())


def chunk_document_text(text: str, chunk_chars: int = MAX_RAG_CHUNK_CHARS) -> list[str]:
    """Split a document into manageable RAG chunks."""
    normalized = re.sub(r"\n{3,}", "\n\n", text).strip()
    if not normalized:
        return []

    paragraphs = [paragraph.strip() for paragraph in normalized.split("\n\n") if paragraph.strip()]
    chunks: list[str] = []
    current = ""

    for paragraph in paragraphs:
        if len(paragraph) > chunk_chars:
            for index in range(0, len(paragraph), chunk_chars):
                piece = paragraph[index:index + chunk_chars].strip()
                if piece:
                    if current:
                        chunks.append(current)
                        current = ""
                    chunks.append(piece)
            continue

        candidate = paragraph if not current else f"{current}\n\n{paragraph}"
        if len(candidate) <= chunk_chars:
            current = candidate
        else:
            if current:
                chunks.append(current)
            current = paragraph

    if current:
        chunks.append(current)

    return chunks[:MAX_RAG_CHUNKS]


def collect_webdav_documents(
    base_url: str,
    username: str,
    password: str,
    read_paths: list[str],
    subdir_path: str,
) -> tuple[list[dict], list[str]]:
    """Fetch markdown and PDF files from configured WebDAV paths."""
    session = requests.Session()
    session.auth = (username, password)

    normalized_base_url = normalize_webdav_base_url(base_url)
    documents: list[dict] = []
    status_lines: list[str] = []
    visited_directories: set[str] = set()
    seen_files: set[str] = set()

    for configured_path in [path.strip() for path in read_paths if path.strip()]:
        scan_root = join_webdav_paths(configured_path, subdir_path)
        pending_urls = [build_webdav_url(normalized_base_url, scan_root)]

        while pending_urls and len(documents) < MAX_RAG_FILES:
            current_url = pending_urls.pop(0)
            if current_url in visited_directories:
                continue
            visited_directories.add(current_url)

            listing_root = webdav_propfind(session, current_url, depth="1")
            entries = parse_webdav_listing(listing_root, scan_root)
            current_relative = urlparse(unquote(current_url)).path.removeprefix(scan_root).strip("/")

            for entry in entries:
                relative_path = entry["relative_path"]
                if relative_path == current_relative:
                    continue

                if entry["is_collection"]:
                    pending_urls.append(build_webdav_url(normalized_base_url, entry["full_path"]))
                    continue

                suffix = Path(relative_path).suffix.lower()
                if suffix not in {".md", ".markdown", ".pdf"}:
                    continue
                if relative_path in seen_files:
                    continue

                file_response = session.get(build_webdav_url(normalized_base_url, entry["full_path"]), timeout=WEBDAV_TIMEOUT)
                file_response.raise_for_status()
                seen_files.add(relative_path)

                if suffix == ".pdf":
                    text = extract_pdf_text(file_response.content)
                else:
                    text = file_response.content.decode("utf-8", errors="replace")

                if not text.strip():
                    status_lines.append(f"Skipped empty document: {relative_path}")
                    continue

                documents.append(
                    {
                        "path": relative_path,
                        "type": suffix.lstrip("."),
                        "size": entry.get("size"),
                        "text": text,
                    }
                )
                status_lines.append(f"Loaded {relative_path}")

    return documents, status_lines


def build_rag_index(documents: list[dict]) -> list[dict]:
    """Create searchable text chunks from WebDAV documents."""
    chunks: list[dict] = []
    for document in documents:
        for index, chunk_text in enumerate(chunk_document_text(document["text"]), start=1):
            chunks.append(
                {
                    "path": document["path"],
                    "type": document["type"],
                    "chunk_index": index,
                    "text": chunk_text,
                    "tokens": set(tokenize_query(chunk_text)),
                }
            )
            if len(chunks) >= MAX_RAG_CHUNKS:
                return chunks
    return chunks


def select_rag_chunks(query: str, rag_chunks: list[dict], top_k: int = RAG_TOP_K) -> list[dict]:
    """Return the most relevant chunks for the current query."""
    query_tokens = tokenize_query(query)
    if not query_tokens:
        return rag_chunks[:top_k]

    scored_chunks: list[tuple[int, int, dict]] = []
    unique_query_tokens = set(query_tokens)
    for chunk in rag_chunks:
        overlap = len(unique_query_tokens.intersection(chunk["tokens"]))
        if overlap <= 0:
            continue
        scored_chunks.append((overlap, len(chunk["text"]), chunk))

    scored_chunks.sort(key=lambda item: (-item[0], item[1]))
    if scored_chunks:
        return [item[2] for item in scored_chunks[:top_k]]
    return rag_chunks[: min(top_k, len(rag_chunks))]


def build_rag_context(query: str, rag_chunks: list[dict]) -> str:
    """Format selected RAG chunks into prompt context."""
    selected_chunks = select_rag_chunks(query, rag_chunks)
    if not selected_chunks:
        return ""

    sections = []
    for chunk in selected_chunks:
        sections.append(
            "\n".join(
                [
                    f"[Source] {chunk['path']}",
                    f"[Chunk] {chunk['chunk_index']}",
                    chunk["text"].strip(),
                ]
            )
        )
    return "\n\n".join(sections)


def image_to_base64(uploaded_file) -> tuple[str, Image.Image]:
    """Return a base64 string for Ollama and a PIL preview image."""
    data = uploaded_file.getvalue()
    encoded = base64.b64encode(data).decode("utf-8")
    image = Image.open(io.BytesIO(data))
    return encoded, image


def save_uploaded_excel(uploaded_file) -> Path:
    """Persist an uploaded Excel file to the workspace directory."""
    WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)
    destination = WORKSPACE_DIR / Path(uploaded_file.name).name
    destination.write_bytes(uploaded_file.getvalue())
    return destination


def safe_excel_path(relative_path: str) -> Path:
    """Resolve a workspace path and require an Excel file extension."""
    path = safe_workspace_path(relative_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
        raise ValueError(f"Not an Excel file: {relative_path}")
    return path


def safe_workspace_path(relative_path: str) -> Path:
    """Resolve a user path while preventing access outside the workspace."""
    raw_path = relative_path.strip() or "."
    candidate = (WORKSPACE_DIR / raw_path).resolve()
    workspace_root = WORKSPACE_DIR.resolve()
    if candidate != workspace_root and workspace_root not in candidate.parents:
        raise ValueError("Access outside the workspace directory is not allowed.")
    return candidate


def workspace_relative(path: Path) -> str:
    """Return a workspace-relative display path."""
    return str(path.resolve().relative_to(WORKSPACE_DIR.resolve()))


def list_workspace_files(limit: int = 200) -> list[Path]:
    """Return workspace files for UI download actions."""
    if not WORKSPACE_DIR.exists():
        return []
    files = sorted(path for path in WORKSPACE_DIR.rglob("*") if path.is_file())
    return files[:limit]


def list_workspace_entries(relative_path: str = ".", recursive: bool = False) -> str:
    """List files and directories within the workspace."""
    path = safe_workspace_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"Path not found: {relative_path}")
    if not path.is_dir():
        raise NotADirectoryError(f"Not a directory: {relative_path}")

    if recursive:
        entries = sorted(path.rglob("*"))
    else:
        entries = sorted(path.iterdir())

    lines = []
    for entry in entries[:200]:
        entry_type = "dir" if entry.is_dir() else "file"
        lines.append(f"{entry_type}: {workspace_relative(entry)}")
    if len(entries) > 200:
        lines.append(f"... truncated, total entries: {len(entries)}")
    return "\n".join(lines) if lines else "(empty directory)"


def read_workspace_file(relative_path: str) -> str:
    """Read a text file from the workspace."""
    path = safe_workspace_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {relative_path}")
    if not path.is_file():
        raise IsADirectoryError(f"Path is not a file: {relative_path}")
    if path.stat().st_size > MAX_TOOL_FILE_BYTES:
        raise ValueError(f"File too large to read safely: {relative_path}")
    return path.read_text(encoding="utf-8", errors="replace")


def prompt_requests_workspace_scan(prompt: str) -> bool:
    """Return whether the prompt likely asks the app to inspect workspace files."""
    normalized_prompt = prompt.strip().lower()
    if not normalized_prompt:
        return False
    if "작업파일" in normalized_prompt:
        return True

    command_patterns = [
        r"파일\s*내용.*알려",
        r"파일\s*내용.*보여",
        r"파일\s*내용.*확인",
        r"파일.*읽어",
        r"workspace.*파일.*알려",
        r"workspace.*찾아",
        r"look at .*file",
        r"read .*file",
        r"find .*file",
        r"check .*file",
    ]
    if any(re.search(pattern, normalized_prompt) for pattern in command_patterns):
        return True
    return bool(extract_workspace_file_references(prompt))


def extract_workspace_file_references(prompt: str) -> list[str]:
    """Extract likely file names or relative paths mentioned in the prompt."""
    references: list[str] = []
    extension_pattern = r"(?:txt|md|markdown|csv|json|yaml|yml|log|py|js|ts|tsx|jsx|html|css|xml|ini|conf|cfg|sh)"
    patterns = [
        rf"[\w./-]+\.{extension_pattern}",
        rf"[\w-]+\s*\.{extension_pattern}",
    ]
    for pattern in patterns:
        for match in re.findall(pattern, prompt, flags=re.IGNORECASE):
            if isinstance(match, tuple):
                continue
            candidate = re.sub(r"\s+", "", str(match).strip().strip("\"'`"))
            if candidate and candidate not in references:
                references.append(candidate)
    return references


def extract_workspace_excel_references(prompt: str) -> list[str]:
    """Extract likely Excel file names or relative paths mentioned in the prompt."""
    references: list[str] = []
    patterns = [
        r"[\w./-]+\.(?:xlsx|xlsm|xltx|xltm|xls)",
        r"[\w-]+\s*\.(?:xlsx|xlsm|xltx|xltm|xls)",
    ]
    for pattern in patterns:
        for match in re.findall(pattern, prompt, flags=re.IGNORECASE):
            candidate = re.sub(r"\s+", "", str(match).strip().strip("\"'`"))
            if candidate and candidate not in references:
                references.append(candidate)
    return references


def extract_task_file_reference(prompt: str) -> str:
    """Extract the file reference that follows the '작업파일' marker."""
    quoted_match = re.search(
        r"작업파일\s*[:=]?\s*[\"'`](?P<path>.+?)[\"'`]",
        prompt,
        flags=re.IGNORECASE,
    )
    if quoted_match:
        return quoted_match.group("path").strip()

    bare_match = re.search(
        r"작업파일\s*[:=]?\s*(?P<path>[^\s,;]+)",
        prompt,
        flags=re.IGNORECASE,
    )
    if not bare_match:
        return ""

    candidate = bare_match.group("path").strip().strip("\"'`")
    trailing_noise = {"읽어줘", "읽기", "확인", "처리", "분석", "보기", "알려줘", "열어줘"}
    if candidate in trailing_noise:
        return ""
    return candidate


def resolve_prompt_file_references(prompt: str, workspace_files: list[Path]) -> list[Path]:
    """Resolve prompt-mentioned file names or relative paths to workspace files."""
    references = extract_workspace_file_references(prompt)
    if not references:
        return []

    resolved_files: list[Path] = []
    for reference in references:
        normalized_reference = reference.replace("\\", "/").strip().lower().lstrip("./")
        basename_reference = Path(normalized_reference).name
        for path in workspace_files:
            try:
                relative_path = workspace_relative(path).replace("\\", "/").lower()
            except OSError:
                continue
            if relative_path == normalized_reference or relative_path.endswith(f"/{normalized_reference}"):
                if path not in resolved_files:
                    resolved_files.append(path)
                continue
            if basename_reference and Path(relative_path).name == basename_reference:
                if path not in resolved_files:
                    resolved_files.append(path)
    return resolved_files


def resolve_workspace_file_reference(reference: str, workspace_files: list[Path]) -> Path | None:
    """Resolve one prompt-provided file name or relative path to a workspace file."""
    if not reference:
        return None

    normalized_reference = reference.replace("\\", "/").strip().lower().lstrip("./")
    basename_reference = Path(normalized_reference).name
    stem_reference = Path(basename_reference).stem or basename_reference

    basename_matches: list[Path] = []
    stem_matches: list[Path] = []
    for path in workspace_files:
        try:
            relative_path = workspace_relative(path).replace("\\", "/").lower()
        except OSError:
            continue
        if relative_path == normalized_reference or relative_path.endswith(f"/{normalized_reference}"):
            return path
        if basename_reference and Path(relative_path).name == basename_reference:
            basename_matches.append(path)
            continue
        if stem_reference and Path(relative_path).stem.lower() == stem_reference.lower():
            stem_matches.append(path)

    if basename_matches:
        basename_matches.sort(key=lambda path: len(workspace_relative(path)))
        return basename_matches[0]
    if stem_matches:
        stem_matches.sort(key=lambda path: len(workspace_relative(path)))
        return stem_matches[0]
    return None


def prompt_requests_excel_statistics(prompt: str) -> bool:
    """Return whether the prompt likely asks for Excel statistics."""
    normalized_prompt = prompt.strip().lower()
    if not normalized_prompt:
        return False

    statistic_keywords = [
        "통계",
        "평균",
        "합계",
        "총합",
        "최대",
        "최소",
        "건수",
        "개수",
        "중앙값",
        "표준편차",
        "average",
        "mean",
        "sum",
        "total",
        "max",
        "min",
        "count",
        "median",
        "std",
        "statistics",
    ]
    return bool(extract_workspace_excel_references(prompt)) and any(
        keyword in normalized_prompt for keyword in statistic_keywords
    )


def build_task_file_context(prompt: str, max_chars_per_file: int = MAX_WORKSPACE_SCAN_CHARS) -> tuple[str, str, bool]:
    """Open the file specified by '작업파일' from workspace and return context plus status."""
    reference = extract_task_file_reference(prompt)
    if not reference:
        return (
            "The prompt included the 작업파일 marker, but no valid file name was found after it.",
            "작업파일 뒤에서 유효한 파일명을 찾지 못했습니다.",
            "작업파일" in prompt,
        )

    workspace_files = list_workspace_files(limit=500)
    resolved_path = resolve_workspace_file_reference(reference, workspace_files)
    if resolved_path is None:
        return (
            f"The requested task file was not found in the workspace: {reference}",
            f"작업파일 `{reference}` 을(를) workspace에서 찾지 못했습니다.",
            True,
        )

    if resolved_path.stat().st_size > MAX_TOOL_FILE_BYTES:
        return (
            f"The requested task file is too large to open safely: {workspace_relative(resolved_path)}",
            f"작업파일 `{workspace_relative(resolved_path)}` 이(가) 너무 커서 열 수 없습니다.",
            True,
        )

    try:
        with open(resolved_path, "r", encoding="utf-8", errors="replace") as file_handle:
            content = file_handle.read()
    except OSError as exc:
        return (
            f"Failed to open the requested task file: {workspace_relative(resolved_path)} ({exc})",
            f"작업파일 `{workspace_relative(resolved_path)}` 을(를) 열지 못했습니다: {exc}",
            True,
        )

    snippet = content[:max_chars_per_file].strip()
    if len(content) > max_chars_per_file:
        snippet += "\n... [truncated]"
    return (
        f"[Task File] {workspace_relative(resolved_path)}\n{snippet}",
        f"작업파일 `{workspace_relative(resolved_path)}` 을(를) 열었습니다.",
        True,
    )


def is_workspace_text_file(path: Path) -> bool:
    """Return whether the workspace file should be scanned as text."""
    allowed_suffixes = {
        ".txt",
        ".md",
        ".markdown",
        ".csv",
        ".json",
        ".yaml",
        ".yml",
        ".log",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".jsx",
        ".html",
        ".css",
        ".xml",
        ".ini",
        ".conf",
        ".cfg",
        ".sh",
    }
    return path.suffix.lower() in allowed_suffixes


def build_workspace_context_for_prompt(
    prompt: str,
    limit_files: int = MAX_WORKSPACE_SCAN_FILES,
    max_chars_per_file: int = MAX_WORKSPACE_SCAN_CHARS,
) -> tuple[str, int]:
    """Scan workspace files relevant to the prompt and return a prompt context block."""
    workspace_files = list_workspace_files(limit=500)
    if not workspace_files:
        return "", 0

    prompt_terms = {
        token
        for token in re.findall(r"[a-z0-9가-힣._-]+", prompt.lower())
        if len(token) >= 2
    }
    if "workspace" in prompt.lower():
        prompt_terms.add("workspace")

    task_file_context, _, task_file_requested = build_task_file_context(prompt, max_chars_per_file=max_chars_per_file)
    if task_file_requested:
        if task_file_context.startswith("[Task File]"):
            return task_file_context, 1
        return task_file_context, 0

    explicit_files = [
        path
        for path in resolve_prompt_file_references(prompt, workspace_files)
        if is_workspace_text_file(path) and path.stat().st_size <= MAX_TOOL_FILE_BYTES
    ]

    ranked_files: list[tuple[int, Path]] = []
    for path in workspace_files:
        if not is_workspace_text_file(path):
            continue
        try:
            relative_path = workspace_relative(path).lower()
            file_size = path.stat().st_size
        except OSError:
            continue
        if file_size > MAX_TOOL_FILE_BYTES:
            continue

        score = 0
        if path in explicit_files:
            score += 100
        if any(term in relative_path for term in prompt_terms):
            score += 5
        score += sum(1 for term in prompt_terms if term in relative_path)
        if path.suffix.lower() in {".md", ".txt", ".csv", ".json"}:
            score += 1
        ranked_files.append((score, path))

    ranked_files.sort(key=lambda item: (-item[0], str(item[1]).lower()))
    selected_files: list[Path] = []
    for path in explicit_files:
        if path not in selected_files:
            selected_files.append(path)
    for score, path in ranked_files:
        if score <= 0:
            continue
        if path not in selected_files:
            selected_files.append(path)
        if len(selected_files) >= limit_files:
            break
    if not selected_files:
        selected_files = [path for _, path in ranked_files[:limit_files]]
    if not selected_files:
        return "", 0

    context_blocks: list[str] = []
    for path in selected_files:
        try:
            content = read_workspace_file(workspace_relative(path))
        except (OSError, ValueError):
            continue
        snippet = content[:max_chars_per_file].strip()
        if not snippet:
            continue
        if len(content) > max_chars_per_file:
            snippet += "\n... [truncated]"
        context_blocks.append(f"[Workspace File] {workspace_relative(path)}\n{snippet}")

    return "\n\n".join(context_blocks), len(context_blocks)


def write_workspace_file(relative_path: str, content: str) -> str:
    """Write text content to a workspace file."""
    path = safe_workspace_path(relative_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return f"Saved file: {workspace_relative(path)}"


def prepare_workspace_download(relative_path: str) -> str:
    """Register a workspace file for download in the Streamlit UI."""
    path = safe_workspace_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {relative_path}")
    if not path.is_file():
        raise IsADirectoryError(f"Path is not a file: {relative_path}")

    st.session_state.setdefault("prepared_downloads", [])
    prepared_downloads: list[str] = st.session_state["prepared_downloads"]
    normalized_path = workspace_relative(path)
    if normalized_path not in prepared_downloads:
        prepared_downloads.append(normalized_path)

    return (
        f"Prepared download: {normalized_path}. "
        "Tell the user to use the download button shown in the Workspace Downloads section."
    )


def copy_workspace_path(source_path: str, destination_path: str) -> str:
    """Copy a workspace file or directory."""
    source = safe_workspace_path(source_path)
    destination = safe_workspace_path(destination_path)
    if not source.exists():
        raise FileNotFoundError(f"Source not found: {source_path}")
    destination.parent.mkdir(parents=True, exist_ok=True)

    if source.is_dir():
        shutil.copytree(source, destination, dirs_exist_ok=True)
        return f"Copied directory: {workspace_relative(source)} -> {workspace_relative(destination)}"

    shutil.copy2(source, destination)
    return f"Copied file: {workspace_relative(source)} -> {workspace_relative(destination)}"


def delete_workspace_path(relative_path: str) -> str:
    """Delete a workspace file or directory."""
    path = safe_workspace_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"Path not found: {relative_path}")
    if path == WORKSPACE_DIR.resolve():
        raise ValueError("Deleting the workspace root is not allowed.")

    if path.is_dir():
        shutil.rmtree(path)
        return f"Deleted directory: {relative_path}"

    path.unlink()
    return f"Deleted file: {relative_path}"


def excel_workbook_info(relative_path: str) -> str:
    """Return workbook sheet metadata."""
    path = safe_excel_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {relative_path}")

    workbook = load_workbook(path, data_only=False)
    lines = [f"Workbook: {workspace_relative(path)}"]
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines.append(
            f"Sheet: {sheet_name} | rows={sheet.max_row} | cols={sheet.max_column}"
        )
    workbook.close()
    return "\n".join(lines)


def excel_sheet_preview(relative_path: str, sheet_name: str, max_rows: int = 20) -> str:
    """Preview top rows from an Excel sheet."""
    path = safe_excel_path(relative_path)
    preview_rows = max(1, min(int(max_rows), 100))
    frame = pd.read_excel(path, sheet_name=sheet_name).head(preview_rows).fillna("")
    return frame.to_csv(index=False).strip()


def excel_read_cells(relative_path: str, sheet_name: str, cell_range: str) -> str:
    """Read one or more cells from an Excel sheet."""
    path = safe_excel_path(relative_path)
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name]

    if ":" not in cell_range:
        value = sheet[cell_range].value
        workbook.close()
        return f"{cell_range} = {value}"

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    lines = []
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=False,
    ):
        row_items = [f"{cell.coordinate}={cell.value}" for cell in row]
        lines.append(", ".join(row_items))
    workbook.close()
    return "\n".join(lines)


def coerce_excel_value(value: str):
    """Convert text input into a simple Excel cell value."""
    lowered = value.strip().lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return value


def excel_write_cell(relative_path: str, sheet_name: str, cell: str, value: str) -> str:
    """Write a value into a single Excel cell."""
    path = safe_excel_path(relative_path)
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    sheet[cell] = coerce_excel_value(value)
    workbook.save(path)
    workbook.close()
    return f"Updated {workspace_relative(path)} {sheet_name}!{cell} = {value}"


def excel_aggregate_range(relative_path: str, sheet_name: str, cell_range: str, operation: str) -> str:
    """Aggregate numeric values from an Excel range."""
    path = safe_excel_path(relative_path)
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    numbers = []
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        for value in row:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numbers.append(float(value))
    workbook.close()

    if operation not in {"sum", "average", "min", "max", "count"}:
        raise ValueError(f"Unsupported operation: {operation}")
    if operation != "count" and not numbers:
        raise ValueError(f"No numeric values found in range: {cell_range}")

    if operation == "sum":
        result = sum(numbers)
    elif operation == "average":
        result = sum(numbers) / len(numbers)
    elif operation == "min":
        result = min(numbers)
    elif operation == "max":
        result = max(numbers)
    else:
        result = len(numbers)

    return f"{operation}({sheet_name}!{cell_range}) = {result}"


def coerce_filter_value(value):
    """Normalize filter input into a comparable python value."""
    if isinstance(value, str):
        return coerce_excel_value(value)
    return value


def apply_excel_filters(frame: pd.DataFrame, filters: list[dict] | None) -> pd.DataFrame:
    """Apply simple row filters to a dataframe."""
    if not filters:
        return frame

    filtered = frame.copy()
    for filter_rule in filters:
        column = str(filter_rule["column"])
        operator = str(filter_rule["operator"]).lower()
        value = coerce_filter_value(filter_rule["value"])
        if column not in filtered.columns:
            raise ValueError(f"Filter column not found: {column}")

        series = filtered[column]
        if operator == "contains":
            mask = series.astype(str).str.contains(str(value), case=False, na=False)
        else:
            comparable_series = series
            numeric_series = pd.to_numeric(series, errors="coerce")
            if not numeric_series.isna().all() and isinstance(value, (int, float)):
                comparable_series = numeric_series

            if operator == "eq":
                mask = comparable_series == value
            elif operator == "ne":
                mask = comparable_series != value
            elif operator == "gt":
                mask = comparable_series > value
            elif operator == "gte":
                mask = comparable_series >= value
            elif operator == "lt":
                mask = comparable_series < value
            elif operator == "lte":
                mask = comparable_series <= value
            else:
                raise ValueError(f"Unsupported filter operator: {operator}")

        filtered = filtered[mask.fillna(False)]

    return filtered


def sanitize_stat_value(value):
    """Convert pandas/numpy values into JSON-friendly python values."""
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return value
    return value


def excel_calculate_statistics(
    relative_path: str,
    sheet_name: str,
    target_columns: list[str] | None = None,
    group_by: list[str] | None = None,
    filters: list[dict] | None = None,
    statistics: list[str] | None = None,
) -> str:
    """Calculate dataframe-style statistics from an Excel sheet."""
    path = safe_excel_path(relative_path)
    frame = pd.read_excel(path, sheet_name=sheet_name)
    filtered = apply_excel_filters(frame, filters)

    requested_stats = [str(item).lower() for item in (statistics or ["count", "sum", "mean", "min", "max"])]
    supported_stats = {"count", "sum", "mean", "min", "max", "median", "std", "nunique"}
    unsupported = [item for item in requested_stats if item not in supported_stats]
    if unsupported:
        raise ValueError(f"Unsupported statistics requested: {', '.join(unsupported)}")

    group_columns = [str(column) for column in (group_by or [])]
    for column in group_columns:
        if column not in filtered.columns:
            raise ValueError(f"group_by column not found: {column}")

    if target_columns:
        selected_columns = [str(column) for column in target_columns]
        missing_columns = [column for column in selected_columns if column not in filtered.columns]
        if missing_columns:
            raise ValueError(f"Target columns not found: {', '.join(missing_columns)}")
    else:
        numeric_columns = filtered.select_dtypes(include=["number"]).columns.tolist()
        if not numeric_columns:
            numeric_columns = [column for column in filtered.columns if column not in group_columns]
        selected_columns = [str(column) for column in numeric_columns]

    if not selected_columns:
        raise ValueError("No columns available for statistics.")

    working = filtered.copy()
    for column in selected_columns:
        working[column] = pd.to_numeric(working[column], errors="coerce")

    response: dict[str, object] = {
        "workbook": workspace_relative(path),
        "sheet_name": sheet_name,
        "row_count_before_filters": int(len(frame.index)),
        "row_count_after_filters": int(len(filtered.index)),
        "target_columns": selected_columns,
        "group_by": group_columns,
        "statistics": requested_stats,
    }
    if filters:
        response["filters"] = filters

    if group_columns:
        grouped = (
            working.groupby(group_columns, dropna=False)[selected_columns]
            .agg(requested_stats)
            .reset_index()
        )
        grouped.columns = [
            "__".join([str(part) for part in column if str(part)])
            if isinstance(column, tuple)
            else str(column)
            for column in grouped.columns
        ]
        grouped = grouped.where(pd.notna(grouped), None)
        response["grouped_statistics"] = grouped.to_dict(orient="records")
    else:
        per_column_stats: dict[str, dict[str, object]] = {}
        for column in selected_columns:
            column_stats: dict[str, object] = {}
            series = working[column]
            for stat_name in requested_stats:
                if stat_name == "count":
                    value = int(series.count())
                elif stat_name == "nunique":
                    value = int(series.nunique(dropna=True))
                else:
                    value = sanitize_stat_value(getattr(series, stat_name)())
                column_stats[stat_name] = value
            per_column_stats[column] = column_stats
        response["column_statistics"] = per_column_stats

    return json.dumps(response, ensure_ascii=False, indent=2)


def infer_excel_statistics_from_prompt(prompt: str) -> list[str]:
    """Infer requested statistic names from a prompt."""
    normalized_prompt = prompt.lower()
    keyword_map = [
        ("count", ["건수", "개수", "count"]),
        ("sum", ["합계", "총합", "sum", "total"]),
        ("mean", ["평균", "average", "mean"]),
        ("min", ["최소", "min", "minimum"]),
        ("max", ["최대", "max", "maximum"]),
        ("median", ["중앙값", "median"]),
        ("std", ["표준편차", "std", "standard deviation"]),
        ("nunique", ["고유값", "unique", "distinct", "nunique"]),
    ]

    requested: list[str] = []
    for stat_name, keywords in keyword_map:
        if any(keyword in normalized_prompt for keyword in keywords):
            requested.append(stat_name)
    return requested or ["count", "sum", "mean", "min", "max"]


def infer_excel_sheet_name_from_prompt(prompt: str, sheet_names: list[str]) -> str:
    """Infer the most likely sheet name from the prompt."""
    normalized_prompt = prompt.lower()
    for sheet_name in sheet_names:
        if sheet_name.lower() in normalized_prompt:
            return sheet_name
    return sheet_names[0]


def infer_excel_group_by_columns(prompt: str, columns: list[str]) -> list[str]:
    """Infer group-by columns mentioned in the prompt."""
    normalized_prompt = prompt.lower()
    group_columns: list[str] = []
    for column in columns:
        lowered = column.lower()
        if (
            f"{column}별" in prompt
            or f"{lowered}별" in normalized_prompt
            or f"by {lowered}" in normalized_prompt
        ):
            group_columns.append(column)
    return group_columns


def infer_excel_target_columns(prompt: str, columns: list[str], group_columns: list[str]) -> list[str] | None:
    """Infer target columns mentioned in the prompt."""
    normalized_prompt = prompt.lower()
    matched_columns = [
        column for column in columns
        if column not in group_columns and column.lower() in normalized_prompt
    ]
    return matched_columns or None


def infer_excel_filters_from_prompt(prompt: str, columns: list[str]) -> list[dict] | None:
    """Infer simple equality filters from a natural-language prompt."""
    filters: list[dict] = []
    for column in columns:
        escaped_column = re.escape(column)
        patterns = [
            rf"{escaped_column}\s*(?:이|가|은|는)\s*([^\s,]+)\s*인",
            rf"{escaped_column}\s*=\s*([^\s,]+)",
            rf"{escaped_column}\s+is\s+([^\s,]+)",
        ]
        for pattern in patterns:
            match = re.search(pattern, prompt, flags=re.IGNORECASE)
            if match:
                value = match.group(1).strip().strip("\"'`")
                filters.append({"column": column, "operator": "eq", "value": value})
                break
    return filters or None


def build_excel_statistics_context_for_prompt(prompt: str) -> tuple[str, str, bool]:
    """Calculate Excel statistics for Gemma-style prompts and return context plus status."""
    references = extract_workspace_excel_references(prompt)
    if not references:
        return "", "", False

    workspace_files = list_workspace_files(limit=500)
    resolved_path = None
    for reference in references:
        candidate = resolve_workspace_file_reference(reference, workspace_files)
        if candidate is not None and candidate.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
            resolved_path = candidate
            break

    if resolved_path is None:
        return (
            f"The requested Excel file was not found in the workspace: {references[0]}",
            f"엑셀 파일 `{references[0]}` 을(를) workspace에서 찾지 못했습니다.",
            True,
        )

    try:
        excel_file = pd.ExcelFile(resolved_path)
        sheet_name = infer_excel_sheet_name_from_prompt(prompt, excel_file.sheet_names)
        sheet_frame = excel_file.parse(sheet_name)
        columns = [str(column) for column in sheet_frame.columns]
        group_columns = infer_excel_group_by_columns(prompt, columns)
        target_columns = infer_excel_target_columns(prompt, columns, group_columns)
        filters = infer_excel_filters_from_prompt(prompt, columns)
        statistics = infer_excel_statistics_from_prompt(prompt)
        result = excel_calculate_statistics(
            relative_path=workspace_relative(resolved_path),
            sheet_name=sheet_name,
            target_columns=target_columns,
            group_by=group_columns or None,
            filters=filters,
            statistics=statistics,
        )
    except Exception as exc:
        return (
            f"Failed to calculate Excel statistics for {workspace_relative(resolved_path)}: {exc}",
            f"엑셀 통계 계산에 실패했습니다: {exc}",
            True,
        )

    context = (
        f"[Excel Statistics] {workspace_relative(resolved_path)}::{sheet_name}\n"
        f"{result}"
    )
    return (
        context,
        f"엑셀 통계 계산 완료: `{workspace_relative(resolved_path)}` 시트 `{sheet_name}` 기준 결과를 모델에 전달했습니다.",
        True,
    )


def sanitize_sheet_title(title: str, used_titles: set[str]) -> str:
    """Create a valid and unique Excel sheet title."""
    cleaned = "".join(character for character in title if character not in '[]:*?/\\')
    cleaned = cleaned[:31] or "Sheet"
    candidate = cleaned
    index = 1
    while candidate in used_titles:
        suffix = f"_{index}"
        candidate = f"{cleaned[: 31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate)
    return candidate


def excel_merge_files(source_paths: list[str], output_path: str, mode: str = "append_rows") -> str:
    """Merge multiple Excel files into a single Excel output file."""
    if not source_paths:
        raise ValueError("source_paths must not be empty.")

    output = safe_excel_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    normalized_sources = [safe_excel_path(source_path) for source_path in source_paths]
    if output.resolve() in [source.resolve() for source in normalized_sources]:
        raise ValueError("output_path must be different from every source path.")

    if mode == "append_rows":
        frames = []
        for source in normalized_sources:
            excel_file = pd.ExcelFile(source)
            for sheet_name in excel_file.sheet_names:
                frame = excel_file.parse(sheet_name)
                frame["source_file"] = source.name
                frame["source_sheet"] = sheet_name
                frames.append(frame)

        if not frames:
            raise ValueError("No Excel data found to merge.")

        merged = pd.concat(frames, ignore_index=True)
        merged.to_excel(output, index=False)
        return f"Merged {len(normalized_sources)} files into {workspace_relative(output)} using append_rows mode."

    if mode == "separate_sheets":
        used_titles: set[str] = set()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for source in normalized_sources:
                excel_file = pd.ExcelFile(source)
                for sheet_name in excel_file.sheet_names:
                    frame = excel_file.parse(sheet_name)
                    target_sheet = sanitize_sheet_title(f"{source.stem}_{sheet_name}", used_titles)
                    frame.to_excel(writer, sheet_name=target_sheet, index=False)

        return f"Merged {len(normalized_sources)} files into {workspace_relative(output)} using separate_sheets mode."

    raise ValueError(f"Unsupported merge mode: {mode}")


def excel_stack_files_to_single_sheet(
    source_paths: list[str],
    output_path: str,
    sheet_name: str = "MergedData",
    gap_rows: int = 2,
) -> str:
    """Stack multiple Excel file contents into a single worksheet vertically."""
    if not source_paths:
        raise ValueError("source_paths must not be empty.")

    output = safe_excel_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    normalized_sources = [safe_excel_path(source_path) for source_path in source_paths]
    if output.resolve() in [source.resolve() for source in normalized_sources]:
        raise ValueError("output_path must be different from every source path.")

    gap_rows = max(0, int(gap_rows))
    output_sheet_name = sanitize_sheet_title(sheet_name, set())
    current_row = 0

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for source in normalized_sources:
            excel_file = pd.ExcelFile(source)
            for source_sheet_name in excel_file.sheet_names:
                frame = excel_file.parse(source_sheet_name)
                header_frame = pd.DataFrame(
                    [
                        [f"source_file={source.name}"],
                        [f"source_sheet={source_sheet_name}"],
                    ]
                )
                header_frame.to_excel(
                    writer,
                    sheet_name=output_sheet_name,
                    index=False,
                    header=False,
                    startrow=current_row,
                )
                current_row += len(header_frame)

                frame.to_excel(
                    writer,
                    sheet_name=output_sheet_name,
                    index=False,
                    startrow=current_row,
                )
                current_row += len(frame.index) + 1 + gap_rows

    return (
        f"Stacked {len(normalized_sources)} files into single sheet "
        f"{workspace_relative(output)}::{output_sheet_name} with {gap_rows} blank row gap(s)."
    )


def execute_file_tool(name: str, arguments: dict) -> str:
    """Run a validated workspace tool call and return its result."""
    if name == "list_files":
        return list_workspace_entries(
            relative_path=str(arguments.get("path", ".")),
            recursive=bool(arguments.get("recursive", False)),
        )
    if name == "read_file":
        return read_workspace_file(str(arguments["path"]))
    if name == "write_file":
        return write_workspace_file(str(arguments["path"]), str(arguments["content"]))
    if name == "copy_path":
        return copy_workspace_path(
            source_path=str(arguments["source_path"]),
            destination_path=str(arguments["destination_path"]),
        )
    if name == "delete_path":
        return delete_workspace_path(str(arguments["path"]))
    if name == "download_file":
        return prepare_workspace_download(str(arguments["path"]))
    if name == "excel_workbook_info":
        return excel_workbook_info(str(arguments["path"]))
    if name == "excel_sheet_preview":
        return excel_sheet_preview(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            max_rows=int(arguments.get("max_rows", 20)),
        )
    if name == "excel_read_cells":
        return excel_read_cells(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            cell_range=str(arguments["cell_range"]),
        )
    if name == "excel_write_cell":
        return excel_write_cell(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            cell=str(arguments["cell"]),
            value=str(arguments["value"]),
        )
    if name == "excel_aggregate_range":
        return excel_aggregate_range(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            cell_range=str(arguments["cell_range"]),
            operation=str(arguments["operation"]).lower(),
        )
    if name == "excel_calculate_statistics":
        return excel_calculate_statistics(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            target_columns=[str(item) for item in arguments.get("target_columns", [])] or None,
            group_by=[str(item) for item in arguments.get("group_by", [])] or None,
            filters=arguments.get("filters"),
            statistics=[str(item) for item in arguments.get("statistics", [])] or None,
        )
    if name == "excel_merge_files":
        return excel_merge_files(
            source_paths=[str(path) for path in arguments["source_paths"]],
            output_path=str(arguments["output_path"]),
            mode=str(arguments.get("mode", "append_rows")).lower(),
        )
    if name == "excel_stack_files_to_single_sheet":
        return excel_stack_files_to_single_sheet(
            source_paths=[str(path) for path in arguments["source_paths"]],
            output_path=str(arguments["output_path"]),
            sheet_name=str(arguments.get("sheet_name", "MergedData")),
            gap_rows=int(arguments.get("gap_rows", 2)),
        )
    raise ValueError(f"Unknown tool: {name}")


def normalize_tool_arguments(arguments) -> dict:
    """Normalize tool arguments that may arrive as a dict or JSON string."""
    if isinstance(arguments, dict):
        return arguments
    if isinstance(arguments, str):
        parsed = json.loads(arguments)
        if isinstance(parsed, dict):
            return parsed
    raise ValueError(f"Unsupported tool arguments: {arguments}")


def prompt_requests_tool_list(prompt: str) -> bool:
    """Return whether the prompt asks the app to describe available tools."""
    normalized_prompt = prompt.strip().lower()
    if not normalized_prompt:
        return False

    patterns = [
        r"사용\s*가능한\s*tool.*알려",
        r"사용가능한\s*툴.*알려",
        r"tool\s*리스트",
        r"tool\s*목록",
        r"available\s*tools?",
        r"what\s*tools?\s*(can|are)",
    ]
    return any(re.search(pattern, normalized_prompt) for pattern in patterns)


def model_supports_tools(model: str) -> bool:
    """Return whether the selected model should use tool calling."""
    return (
        model.startswith("qwen2.5-coder:")
        or model.startswith("qwen3-coder:")
        or model.startswith("qwen3.5:")
    )


def build_available_tool_response(model: str) -> str:
    """Return a direct answer describing the currently available tools."""
    lines = [f"Current model: `{model}`"]

    if model_supports_tools(model):
        lines.append("Tool calling: enabled")
        lines.append("")
        lines.append("Registered tools:")
        for tool in FILE_TOOLS:
            function = tool.get("function", {})
            tool_name = str(function.get("name", ""))
            description = str(function.get("description", "")).strip()
            lines.append(f"- `{tool_name}`: {description}")
    else:
        lines.append("Tool calling: disabled for this model")
        lines.append("")
        lines.append("Available app-side helper actions:")
        lines.append("- `작업파일 <file>`: find the file in `workspace`, open it with Python `open()`, and pass the content to the model")
        lines.append("- `작업파일: <file>` and `작업파일 \"<file>\"`: same task-file flow with `:` or quoted file names")
        lines.append("- Direct file names such as `config.json`, `notes/todo.md`, or `report.txt`: search the `workspace` and prioritize matched files")
        lines.append("- File-reading prompts such as `파일 내용 알려줘`, `파일 읽어`, `read file`, `find file`, `check file`: scan text-like files in `workspace` and pass excerpts to the model")
        lines.append("- Excel statistics prompts such as `sales.xlsx 의 Sheet1 에서 amount 평균 계산해줘`: read the Excel sheet, calculate statistics in the app, and pass the result to the model")
        lines.append("")
        lines.append("If you want true Ollama tool calling, switch to `qwen2.5-coder:7b`, `qwen3-coder:30b`, or `qwen3.5:9b`.")

    return "\n".join(lines)


def build_user_message(
    prompt: str,
    excel_contexts: Iterable[str],
    rag_context: str,
    workspace_context: str,
    allow_tools: bool,
) -> str:
    """Build the final prompt content sent to the model."""
    content_parts = [prompt.strip()]

    excel_contexts = [context for context in excel_contexts if context.strip()]
    if excel_contexts:
        content_parts.append(
            "The user also uploaded Excel data. Use the workbook summaries below when relevant.\n\n"
            + "\n\n".join(excel_contexts)
        )
    if rag_context.strip():
        content_parts.append(
            "The user also connected WebDAV RAG sources. Use the retrieved context below when relevant, and cite the source path in your answer when you rely on it.\n\n"
            + rag_context.strip()
        )
    if workspace_context.strip():
        content_parts.append(
            "The app scanned the workspace for relevant files because the user asked about file contents. "
            "Use the file excerpts below when relevant, and mention the workspace file path when you rely on it.\n\n"
            + workspace_context.strip()
        )
    if allow_tools:
        content_parts.append(
            "You may use workspace tools when needed. "
            "If the user asks to download a workspace file, call download_file for that file so the UI can expose a download button. "
            f"The workspace root is: {WORKSPACE_DIR.resolve()}"
        )

    return "\n\n".join(part for part in content_parts if part)


def get_model_temperature(model: str) -> float:
    """Return the selected or default temperature for a model."""
    model_temperatures = st.session_state.get("model_temperatures", {})
    if model in model_temperatures:
        return float(model_temperatures[model])
    return float(MODEL_DEFAULT_TEMPERATURES.get(model, 0.7))


def summarize_tool_calls(tool_calls: list[dict]) -> str:
    """Return a stable summary string for loop detection and diagnostics."""
    summary_parts: list[str] = []
    for tool_call in tool_calls:
        function = tool_call.get("function", {})
        tool_name = str(function.get("name", ""))
        arguments = normalize_tool_arguments(function.get("arguments", {}) or {})
        normalized_arguments = json.dumps(arguments, ensure_ascii=True, sort_keys=True)
        summary_parts.append(f"{tool_name}:{normalized_arguments}")
    return " | ".join(summary_parts)


def request_final_answer(
    url: str,
    model: str,
    messages: list[dict],
    temperature: float,
    reason: str,
) -> str:
    """Ask the model for a final answer without exposing more tool options."""
    final_messages = list(messages)
    final_messages.append(
        {
            "role": "system",
            "content": (
                "Stop calling tools now. "
                f"Reason: {reason} "
                "Provide the best final answer using only the information already available in this conversation. "
                "If the task could not be completed fully, state what remains."
            ),
        }
    )
    response = requests.post(
        url,
        json={
            "model": model,
            "stream": False,
            "messages": final_messages,
            "options": {
                "temperature": temperature,
            },
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    event = response.json()
    return event.get("message", {}).get("content", "").strip()


def call_ollama(
    host: str,
    model: str,
    prompt: str,
    excel_contexts: list[str],
    rag_context: str,
    workspace_context: str,
    images: list[str],
    temperature: float,
) -> str:
    """Send a chat request to Ollama and allow validated workspace tool calls."""
    url = f"{host.rstrip('/')}/api/chat"
    allow_tools = model_supports_tools(model)
    messages = [
        {
            "role": "system",
            "content": (
                "You are a helpful assistant running on an RTX 5090 server. "
                "Answer clearly and use uploaded Excel context when provided. "
                + (
                    "Analyze images when attached, use workspace file tools when they are needed to complete the user's request, "
                    "and never claim you changed files unless a tool call succeeded."
                    if allow_tools
                    else "Analyze images when attached. Do not claim tool usage or file changes."
                )
            ),
        },
        {
            "role": "user",
            "content": build_user_message(prompt, excel_contexts, rag_context, workspace_context, allow_tools),
            "images": images,
        },
    ]

    if not allow_tools:
        response = requests.post(
            url,
            json={
                "model": model,
                "stream": False,
                "messages": messages,
                "options": {
                    "temperature": temperature,
                },
            },
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        event = response.json()
        return event.get("message", {}).get("content", "").strip()

    previous_tool_signature = ""
    repeated_tool_rounds = 0

    for _ in range(MAX_TOOL_ROUNDS):
        if st.session_state.get("query_cancel_requested"):
            raise RuntimeError("Query stopped by user.")

        response = requests.post(
            url,
            json={
                "model": model,
                "stream": False,
                "messages": messages,
                "tools": FILE_TOOLS,
                "options": {
                    "temperature": temperature,
                },
            },
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        event = response.json()
        message = event.get("message", {})
        messages.append(message)
        tool_calls = message.get("tool_calls") or []

        if not tool_calls:
            return message.get("content", "").strip()

        tool_signature = summarize_tool_calls(tool_calls)
        if tool_signature == previous_tool_signature:
            repeated_tool_rounds += 1
        else:
            repeated_tool_rounds = 0
            previous_tool_signature = tool_signature

        if repeated_tool_rounds >= MAX_IDENTICAL_TOOL_ROUNDS:
            return request_final_answer(
                url=url,
                model=model,
                messages=messages,
                temperature=temperature,
                reason=(
                    "the model repeated the same tool call sequence multiple times "
                    f"({tool_signature})"
                ),
            )

        for tool_call in tool_calls:
            if st.session_state.get("query_cancel_requested"):
                raise RuntimeError("Query stopped by user.")

            function = tool_call.get("function", {})
            tool_name = function.get("name", "")
            arguments = normalize_tool_arguments(function.get("arguments", {}) or {})

            try:
                result = execute_file_tool(tool_name, arguments)
            except Exception as exc:
                result = f"Tool error: {exc}"

            messages.append(
                {
                    "role": "tool",
                    "tool_name": tool_name,
                    "content": str(result),
                }
            )

    final_answer = request_final_answer(
        url=url,
        model=model,
        messages=messages,
        temperature=temperature,
        reason=(
            "the maximum tool round limit was reached "
            f"after {MAX_TOOL_ROUNDS} rounds"
        ),
    )
    if final_answer:
        return final_answer
    raise RuntimeError(
        "Tool calling stopped after reaching the maximum tool round limit."
        f" Rounds attempted: {MAX_TOOL_ROUNDS}."
    )


def fetch_installed_models(host: str) -> dict[str, dict]:
    """Return installed Ollama models keyed by name."""
    response = requests.get(f"{host.rstrip('/')}/api/tags", timeout=30)
    response.raise_for_status()
    data = response.json()
    models = data.get("models", [])
    return {
        model.get("name", "").strip(): model
        for model in models
        if model.get("name", "").strip()
    }


def fetch_model_show_info(host: str, model: str) -> dict:
    """Return detailed model information from Ollama show API."""
    response = requests.post(
        f"{host.rstrip('/')}/api/show",
        json={"model": model},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def detect_gpu_info() -> tuple[str | None, int | None, str | None]:
    """Detect GPU name and memory in GiB using nvidia-smi when available."""
    if not shutil.which("nvidia-smi"):
        return None, None, "nvidia-smi not found"

    try:
        result = subprocess.run(
            [
                "nvidia-smi",
                "--query-gpu=name,memory.total",
                "--format=csv,noheader,nounits",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
    except (subprocess.CalledProcessError, OSError) as exc:
        return None, None, str(exc)

    first_line = result.stdout.strip().splitlines()[0] if result.stdout.strip() else ""
    if not first_line:
        return None, None, "No GPU information returned"

    parts = [part.strip() for part in first_line.split(",", maxsplit=1)]
    if len(parts) != 2:
        return None, None, f"Unexpected nvidia-smi output: {first_line}"

    gpu_name = parts[0]
    try:
        memory_mib = int(parts[1])
    except ValueError:
        return gpu_name, None, f"Unexpected memory value: {parts[1]}"

    memory_gib = max(memory_mib // 1024, 1)
    return gpu_name, memory_gib, None


def recommend_models_for_gpu(memory_gib: int | None) -> tuple[str | None, list[str]]:
    """Return a recommended default model and compatible model list."""
    if memory_gib is None:
        return None, []

    compatible_models = [
        model_name
        for model_name, required_gb in MODEL_MEMORY_GUIDE_GB.items()
        if memory_gib >= required_gb
    ]

    if not compatible_models:
        return "gemma3:1b", ["gemma3:1b"]

    return compatible_models[-1], compatible_models


def format_bytes(size_bytes: int | None) -> str:
    """Format a byte count into a readable string."""
    if size_bytes is None:
        return "Unknown"

    value = float(size_bytes)
    units = ["B", "KB", "MB", "GB", "TB"]
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.2f} {unit}"
        value /= 1024
    return f"{size_bytes} B"


def format_duration(seconds: float | None) -> str:
    """Format seconds into a short readable duration."""
    if seconds is None or seconds < 0:
        return "Unknown"

    total_seconds = int(seconds)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)

    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def render_workspace_downloads() -> None:
    """Render browser download controls for files stored in the workspace."""
    st.subheader("Workspace Files")
    workspace_files = list_workspace_files()

    if not workspace_files:
        st.caption("No files are currently stored in the workspace.")
        return

    st.caption("Download files created on the server workspace directly to this browser client.")
    for workspace_file in workspace_files:
        relative_path = workspace_relative(workspace_file)
        file_size = format_bytes(workspace_file.stat().st_size)
        confirm_key = f"confirm-delete-{relative_path}"
        if confirm_key not in st.session_state:
            st.session_state[confirm_key] = False

        name_col, size_col, download_col, delete_col = st.columns([3.6, 1.0, 1.4, 1.4])
        with name_col:
            st.write(f"`{relative_path}`")
        with size_col:
            st.caption(file_size)
        with download_col:
            st.download_button(
                "Download",
                data=workspace_file.read_bytes(),
                file_name=workspace_file.name,
                mime="application/octet-stream",
                key=f"download-{relative_path}",
                use_container_width=True,
            )
        with delete_col:
            delete_label = "Confirm Delete" if st.session_state[confirm_key] else "Delete"
            if st.button(delete_label, key=f"delete-{relative_path}", use_container_width=True):
                if st.session_state[confirm_key]:
                    workspace_file.unlink()
                    st.session_state[confirm_key] = False
                    st.session_state["workspace_file_message"] = f"Deleted: {relative_path}"
                    st.rerun()
                else:
                    st.session_state[confirm_key] = True

    workspace_message = st.session_state.get("workspace_file_message", "")
    if workspace_message:
        st.success(workspace_message)
        st.session_state["workspace_file_message"] = ""


def get_default_ollama_models_root() -> Path:
    """Return the configured or default local Ollama model root."""
    configured = os.getenv("OLLAMA_MODELS")
    if configured:
        return Path(configured).expanduser()
    return Path.home() / ".ollama" / "models"


def get_selected_ollama_models_root() -> Path:
    """Return the user-selected model root or the default root."""
    selected = st.session_state.get("ollama_models_path")
    if selected:
        return Path(selected).expanduser()
    return get_default_ollama_models_root()


def move_model_storage(source_root: Path, destination_root: Path) -> str:
    """Move existing Ollama model files to a new storage root."""
    source = source_root.expanduser().resolve()
    destination = destination_root.expanduser().resolve()

    if source == destination:
        return f"Model storage path is already set to: {destination}"

    if not source.exists():
        destination.mkdir(parents=True, exist_ok=True)
        return f"Source model path does not exist yet. Created destination: {destination}"

    destination.mkdir(parents=True, exist_ok=True)
    moved_items = []
    for child in source.iterdir():
        target = destination / child.name
        shutil.move(str(child), str(target))
        moved_items.append(child.name)

    return (
        f"Moved {len(moved_items)} item(s) from {source} to {destination}. "
        "Restart Ollama with OLLAMA_MODELS pointing to the new path to use it for future downloads."
    )


def get_model_storage_path(model: str) -> Path:
    """Infer the local Ollama manifest path for a model tag."""
    registry = "registry.ollama.ai"
    namespace = "library"
    repository = model
    tag = "latest"

    if ":" in model:
        repository, tag = model.rsplit(":", maxsplit=1)

    if "/" in repository:
        namespace, repository = repository.split("/", maxsplit=1)

    return get_selected_ollama_models_root() / "manifests" / registry / namespace / repository / tag


def get_selected_model_info(host: str, model: str, installed_model_map: dict[str, dict]) -> tuple[dict | None, str | None]:
    """Return selected model metadata from tags and show API."""
    installed_info = installed_model_map.get(model)
    try:
        show_info = fetch_model_show_info(host, model)
    except requests.RequestException as exc:
        if installed_info:
            return {
                "installed": installed_info,
                "show": {},
            }, str(exc)
        return None, str(exc)
    return {
        "installed": installed_info,
        "show": show_info,
    }, None


def pull_model(host: str, model: str, status_placeholder, progress_placeholder) -> None:
    """Download a model from Ollama and show progress in the sidebar."""
    response = requests.post(
        f"{host.rstrip('/')}/api/pull",
        json={"model": model, "stream": True},
        timeout=REQUEST_TIMEOUT,
        stream=True,
    )
    response.raise_for_status()

    progress_bar = progress_placeholder.progress(0)
    last_status = "Starting download..."
    status_placeholder.info(last_status)
    started_at = time.perf_counter()

    for raw_line in response.iter_lines():
        if not raw_line:
            continue

        event = json.loads(raw_line.decode("utf-8"))
        status_text = event.get("status", last_status)
        total = event.get("total")
        completed = event.get("completed")

        if total and completed is not None and total > 0:
            ratio = min(max(completed / total, 0.0), 1.0)
            elapsed_seconds = max(time.perf_counter() - started_at, 0.001)
            bytes_per_second = completed / elapsed_seconds if completed > 0 else 0.0
            remaining_bytes = max(total - completed, 0)
            eta_seconds = remaining_bytes / bytes_per_second if bytes_per_second > 0 else None
            progress_text = (
                f"{status_text} | {ratio * 100:.1f}% | "
                f"{format_bytes(completed)} / {format_bytes(total)} | "
                f"ETA {format_duration(eta_seconds)}"
            )
            progress_bar.progress(ratio, text=progress_text)
        else:
            progress_bar.progress(0, text=status_text)

        last_status = status_text
        status_placeholder.info(status_text)

    progress_bar.progress(1.0, text="Download complete")
    status_placeholder.success(f"Model download completed: {model}")


def render_sidebar() -> tuple[str, str]:
    st.sidebar.header("Runtime")
    host = st.sidebar.text_input("Ollama Host", value=DEFAULT_OLLAMA_HOST)
    st.sidebar.write(f"Workspace: `{WORKSPACE_DIR.resolve()}`")

    default_model_root = get_default_ollama_models_root()
    saved_model_root = get_saved_model_root()
    if "ollama_models_path" not in st.session_state:
        st.session_state.ollama_models_path = saved_model_root or str(default_model_root)
    if "model_path_message" not in st.session_state:
        st.session_state.model_path_message = ""

    model_root_input = st.sidebar.text_input(
        "Model Download Path",
        value=st.session_state.ollama_models_path,
        help="This path is used for display and migration. Ollama must be restarted with OLLAMA_MODELS set to this path for future downloads to use it.",
    )
    apply_model_root = st.sidebar.button("Use Model Path", use_container_width=True)
    move_model_root = st.sidebar.button("Move Existing Model Files", use_container_width=True)

    if apply_model_root:
        st.session_state.ollama_models_path = persist_model_root(
            model_root_input.strip() or str(default_model_root)
        )
        st.session_state.model_path_message = (
            "Model path updated and saved. Restart Ollama with OLLAMA_MODELS set to this path for future downloads."
        )
        st.rerun()

    if move_model_root:
        try:
            source_root = get_selected_ollama_models_root()
            destination_root = Path(model_root_input.strip() or str(default_model_root))
            message = move_model_storage(source_root, destination_root)
            st.session_state.ollama_models_path = persist_model_root(str(destination_root.expanduser()))
            st.session_state.model_path_message = message
            st.rerun()
        except Exception as exc:
            st.session_state.model_path_message = f"Model file move failed: {exc}"

    if st.session_state.model_path_message:
        st.sidebar.info(st.session_state.model_path_message)

    refresh_models = st.sidebar.button("Refresh Installed Models", use_container_width=True)

    if refresh_models or "installed_models" not in st.session_state:
        try:
            st.session_state.installed_models = fetch_installed_models(host)
            st.session_state.model_error = ""
        except requests.RequestException as exc:
            st.session_state.installed_models = {}
            st.session_state.model_error = f"Failed to load installed models: {exc}"

    installed_model_map = st.session_state.get("installed_models", {})
    installed_models = sorted(installed_model_map.keys())
    model_candidates = list(dict.fromkeys(SUPPORTED_MODEL_OPTIONS + installed_models))
    gpu_name, gpu_memory_gib, gpu_error = detect_gpu_info()
    recommended_model, compatible_models = recommend_models_for_gpu(gpu_memory_gib)

    if "preferred_model" not in st.session_state:
        st.session_state.preferred_model = recommended_model or DEFAULT_MODEL

    if "auto_select_downloaded_model" not in st.session_state:
        st.session_state.auto_select_downloaded_model = True

    if st.session_state.get("selected_model") not in model_candidates:
        preferred = st.session_state.get("preferred_model", DEFAULT_MODEL)
        if preferred in model_candidates:
            st.session_state.selected_model = preferred
        elif DEFAULT_MODEL in model_candidates:
            st.session_state.selected_model = DEFAULT_MODEL
        else:
            st.session_state.selected_model = model_candidates[0]

    st.sidebar.subheader("GPU Recommendation")
    if gpu_name and gpu_memory_gib:
        st.sidebar.write(f"GPU: `{gpu_name}`")
        st.sidebar.write(f"Memory: `{gpu_memory_gib} GiB`")
        if recommended_model:
            st.sidebar.success(f"Recommended default: `{recommended_model}`")
        if compatible_models:
            st.sidebar.caption("Fits this GPU")
            for compatible_model in compatible_models:
                st.sidebar.code(compatible_model)
    else:
        st.sidebar.info("GPU memory could not be detected automatically.")
        if gpu_error:
            st.sidebar.caption(gpu_error)

    selected_model = st.sidebar.selectbox(
        "Model Select",
        options=model_candidates,
        index=model_candidates.index(st.session_state.selected_model),
    )
    st.session_state.selected_model = selected_model

    custom_model = st.sidebar.text_input("Custom Model Tag", value=selected_model)
    model = custom_model.strip() or selected_model

    if "model_temperatures" not in st.session_state:
        st.session_state.model_temperatures = {}
    current_temperature = get_model_temperature(model)
    selected_temperature = st.sidebar.slider(
        "Model Temperature",
        min_value=0.0,
        max_value=2.0,
        value=float(current_temperature),
        step=0.1,
        help="Temperature is remembered per model and applied to Ollama chat requests.",
    )
    st.session_state.model_temperatures[model] = float(selected_temperature)

    if st.session_state.get("model_error"):
        st.sidebar.warning(st.session_state.model_error)

    with st.sidebar.expander("Installed Models", expanded=False):
        if installed_models:
            for installed_model in installed_models:
                st.write(f"- {installed_model}")
        else:
            st.write("No installed models found.")

    st.session_state.auto_select_downloaded_model = st.sidebar.checkbox(
        "Auto-select downloaded model",
        value=st.session_state.auto_select_downloaded_model,
    )

    if recommended_model and st.sidebar.button("Use Recommended Model", use_container_width=True):
        st.session_state.preferred_model = recommended_model
        st.session_state.selected_model = recommended_model
        st.rerun()

    st.sidebar.markdown("Recommended model options")
    for option in SUPPORTED_MODEL_OPTIONS:
        required_gb = MODEL_MEMORY_GUIDE_GB.get(option)
        label = f"{option}  ({required_gb} GiB+ recommended)" if required_gb else option
        st.sidebar.code(label)

    st.sidebar.subheader("Current Model")
    selected_model_info, selected_model_error = get_selected_model_info(host, model, installed_model_map)
    storage_root = get_selected_ollama_models_root()
    storage_path = get_model_storage_path(model)
    st.sidebar.write(f"Model: `{model}`")
    st.sidebar.write(f"Tool Support: `{'Enabled' if model_supports_tools(model) else 'Disabled'}`")
    st.sidebar.write(f"Temperature: `{get_model_temperature(model):.1f}`")
    st.sidebar.write(f"Storage root: `{storage_root}`")
    st.sidebar.write(f"Storage path: `{storage_path}`")
    st.sidebar.caption("Future downloads use this path only after Ollama is restarted with OLLAMA_MODELS set to the same directory.")

    if selected_model_info:
        installed_info = selected_model_info.get("installed") or {}
        show_info = selected_model_info.get("show") or {}
        details = installed_info.get("details") or {}
        model_size = installed_info.get("size")
        parameter_size = details.get("parameter_size") or show_info.get("details", {}).get("parameter_size")
        quantization = details.get("quantization_level") or show_info.get("details", {}).get("quantization_level")
        family = details.get("family") or show_info.get("details", {}).get("family")

        st.sidebar.write(f"Model size: `{format_bytes(model_size)}`")
        if parameter_size:
            st.sidebar.write(f"Parameters: `{parameter_size}`")
        if quantization:
            st.sidebar.write(f"Quantization: `{quantization}`")
        if family:
            st.sidebar.write(f"Family: `{family}`")
    elif selected_model_error:
        st.sidebar.caption(f"Model detail lookup failed: {selected_model_error}")

    status_placeholder = st.sidebar.empty()
    progress_placeholder = st.sidebar.empty()
    if st.sidebar.button("Download Selected Model", type="primary", use_container_width=True):
        try:
            pull_model(host, model, status_placeholder, progress_placeholder)
            st.session_state.installed_models = fetch_installed_models(host)
            st.session_state.preferred_model = model
            if st.session_state.auto_select_downloaded_model:
                st.session_state.selected_model = model
            st.rerun()
        except requests.RequestException as exc:
            status_placeholder.error(f"Model download failed: {exc}")

    st.sidebar.markdown("Example: `OLLAMA_HOST=http://127.0.0.1:11434 ollama serve`")

    return host, model


def render_prepared_downloads() -> None:
    """Render download buttons for files prepared by tool calls."""
    prepared_downloads = st.session_state.get("prepared_downloads", [])
    if not prepared_downloads:
        return

    st.subheader("Workspace Downloads")
    st.caption("Files prepared by tool calls can be downloaded here.")

    for relative_path in prepared_downloads:
        try:
            path = safe_workspace_path(relative_path)
            if not path.exists() or not path.is_file():
                st.warning(f"Prepared file is no longer available: {relative_path}")
                continue

            data = path.read_bytes()
            st.download_button(
                label=f"Download {relative_path}",
                data=data,
                file_name=path.name,
                mime="application/octet-stream",
                key=f"download::{relative_path}",
                use_container_width=True,
            )
        except Exception as exc:
            st.error(f"Failed to prepare download for {relative_path}: {exc}")


def render_webdav_rag_panel() -> str:
    """Render the right-side WebDAV RAG settings panel and return current RAG context."""
    saved_webdav = get_saved_webdav_settings()
    if "webdav_base_url" not in st.session_state:
        st.session_state.webdav_base_url = saved_webdav["base_url"]
    if "webdav_username" not in st.session_state:
        st.session_state.webdav_username = saved_webdav["username"]
    if "webdav_password" not in st.session_state:
        st.session_state.webdav_password = saved_webdav["password"]
    if "webdav_read_paths" not in st.session_state:
        st.session_state.webdav_read_paths = list(saved_webdav["read_paths"])
    if "webdav_subdir_path" not in st.session_state:
        st.session_state.webdav_subdir_path = saved_webdav["subdir_path"]
    if "webdav_status_message" not in st.session_state:
        st.session_state.webdav_status_message = ""
    if "webdav_documents" not in st.session_state:
        st.session_state.webdav_documents = []
    if "webdav_rag_chunks" not in st.session_state:
        st.session_state.webdav_rag_chunks = []
    if "webdav_rag_enabled" not in st.session_state:
        st.session_state.webdav_rag_enabled = False

    st.subheader("WebDAV / RAG")
    st.caption("Use the server origin as WebDAV Base URL, add one or more WebDAV read roots, and optionally append a shared subdirectory path.")
    st.caption("Password is kept only in this browser's localStorage and is not written into the server app_settings.json file.")
    st.caption("Example: Base URL `https://keties.mooo.com:22443`, Read Path `/remote.php/dav/files/tinyos/`, Subdir `메모`.")
    rag_enabled = st.toggle("RAG Enabled", value=st.session_state.webdav_rag_enabled, key="webdav_rag_enabled_toggle")
    st.session_state.webdav_rag_enabled = rag_enabled

    base_url = st.text_input(
        "WebDAV Base URL",
        value=st.session_state.webdav_base_url,
        placeholder="https://nextcloud.example.com:22443",
        key="webdav_base_url_input",
    )
    username = st.text_input(
        "WebDAV Username",
        value=st.session_state.webdav_username,
        key="webdav_username_input",
    )
    password = st.text_input(
        "WebDAV Password / App Token",
        value=st.session_state.webdav_password,
        type="password",
        key="webdav_password_input",
    )

    read_paths: list[str] = []
    for index in range(4):
        initial_value = st.session_state.webdav_read_paths[index] if index < len(st.session_state.webdav_read_paths) else ""
        read_paths.append(
            st.text_input(
                f"Read Path {index + 1}",
                value=initial_value,
                placeholder=DEFAULT_WEBDAV_READ_PATH_PLACEHOLDER,
                key=f"webdav_read_path_{index + 1}",
            )
        )
    subdir_path = st.text_input(
        "Subdir Path",
        value=st.session_state.webdav_subdir_path,
        placeholder=DEFAULT_WEBDAV_SUBDIR_PLACEHOLDER,
        key="webdav_subdir_path_input",
    )

    render_client_webdav_storage_helper()

    normalized_base_url = normalize_webdav_base_url(base_url)
    normalized_read_paths = [normalize_webdav_read_path(normalized_base_url, path) for path in read_paths]
    normalized_subdir_path = normalize_webdav_subdir_path(subdir_path)
    if any(original.strip() != normalized.strip() for original, normalized in zip(read_paths, normalized_read_paths)):
        st.caption("Read Path values are auto-normalized into server paths.")
    if subdir_path.strip() != normalized_subdir_path:
        st.caption("Subdir Path is auto-normalized as a relative path.")

    save_col, test_col, sync_col = st.columns(3)
    with save_col:
        save_settings = st.button("Save WebDAV Settings", use_container_width=True)
    with test_col:
        test_connection = st.button("Test WebDAV Connection", use_container_width=True)
    with sync_col:
        refresh_rag = st.button("Load WebDAV RAG", type="primary", use_container_width=True)

    st.session_state.webdav_base_url = normalized_base_url.rstrip("/")
    st.session_state.webdav_username = username.strip()
    st.session_state.webdav_password = password
    st.session_state.webdav_read_paths = normalized_read_paths
    st.session_state.webdav_subdir_path = normalized_subdir_path

    if save_settings:
        persist_webdav_settings(
            base_url=st.session_state.webdav_base_url,
            username=st.session_state.webdav_username,
            read_paths=st.session_state.webdav_read_paths,
            subdir_path=st.session_state.webdav_subdir_path,
        )
        st.session_state.webdav_status_message = "WebDAV settings saved. Password stays only in the current browser localStorage."
        st.rerun()

    if test_connection:
        try:
            status_message, status_lines = test_webdav_connection(
                base_url=st.session_state.webdav_base_url,
                username=st.session_state.webdav_username,
                password=st.session_state.webdav_password,
                read_paths=st.session_state.webdav_read_paths,
                subdir_path=st.session_state.webdav_subdir_path,
            )
            st.session_state.webdav_status_message = status_message
            st.session_state.webdav_status_detail = status_lines
            st.rerun()
        except Exception as exc:
            st.session_state.webdav_status_message = f"WebDAV test failed: {exc}"
            st.session_state.webdav_status_detail = []

    if refresh_rag:
        try:
            documents, status_lines = collect_webdav_documents(
                base_url=st.session_state.webdav_base_url,
                username=st.session_state.webdav_username,
                password=st.session_state.webdav_password,
                read_paths=st.session_state.webdav_read_paths,
                subdir_path=st.session_state.webdav_subdir_path,
            )
            st.session_state.webdav_documents = documents
            st.session_state.webdav_rag_chunks = build_rag_index(documents)
            st.session_state.webdav_status_message = (
                f"Loaded {len(documents)} document(s) and {len(st.session_state.webdav_rag_chunks)} chunk(s) from WebDAV."
            )
            st.session_state.webdav_status_detail = status_lines
            persist_webdav_settings(
                base_url=st.session_state.webdav_base_url,
                username=st.session_state.webdav_username,
                read_paths=st.session_state.webdav_read_paths,
                subdir_path=st.session_state.webdav_subdir_path,
            )
            st.rerun()
        except Exception as exc:
            st.session_state.webdav_status_message = f"WebDAV load failed: {exc}"
            st.session_state.webdav_status_detail = []

    if st.session_state.webdav_status_message:
        st.info(st.session_state.webdav_status_message)

    documents: list[dict] = st.session_state.get("webdav_documents", [])
    rag_chunks: list[dict] = st.session_state.get("webdav_rag_chunks", [])
    if documents:
        st.write(f"Documents: `{len(documents)}`")
        st.write(f"Chunks: `{len(rag_chunks)}`")
        with st.expander("Loaded WebDAV Sources", expanded=False):
            for document in documents[:MAX_RAG_FILES]:
                size_label = format_bytes(document.get("size"))
                st.write(f"- `{document['path']}` ({document['type']}, {size_label})")

    status_detail = st.session_state.get("webdav_status_detail", [])
    if status_detail:
        with st.expander("Sync Log", expanded=False):
            for line in status_detail[:100]:
                st.write(f"- {line}")

    if not st.session_state.webdav_rag_enabled:
        st.warning("RAG is off. The current query will run without WebDAV context.")
        return ""

    current_prompt = st.session_state.get("prompt_input", "")
    return build_rag_context(current_prompt, rag_chunks)


def render_history_item(item: dict, fallback_model: str, fallback_temperature: float) -> None:
    """Render one query/answer history item with visual separation."""
    prompt_text = html.escape(str(item.get("prompt", "")))
    answer_text = html.escape(str(item.get("answer", "")))
    queried_at = item.get("queried_at", "-")
    elapsed_seconds = item.get("elapsed_seconds")
    elapsed_label = "-" if elapsed_seconds is None else f"{elapsed_seconds:.2f} seconds"
    model_label = item.get("model", fallback_model)
    temperature_label = item.get("temperature", fallback_temperature)

    st.markdown(
        f"""
        <div style="border: 1px solid #d7e3f4; border-radius: 12px; overflow: hidden; margin-bottom: 1rem;">
            <div style="background: #eef6ff; padding: 0.85rem 1rem; border-bottom: 1px solid #d7e3f4;">
            <div style="font-weight: 700; color: #0f3d75; margin-bottom: 0.35rem;">Prompt</div>
            <div style="white-space: pre-wrap; color: #122033;">{prompt_text}</div>
          </div>
          <div style="background: #fff8e8; padding: 0.85rem 1rem; border-bottom: 1px solid #f0dfb0;">
            <div style="font-weight: 700; color: #7a4b00; margin-bottom: 0.35rem;">Answer</div>
            <div style="white-space: pre-wrap; color: #2a1d00;">{answer_text}</div>
          </div>
          <div style="background: #f7f9fc; padding: 0.65rem 1rem; color: #425466; font-size: 0.92rem;">
            Query time: {queried_at} | Model: {model_label} | Temperature: {temperature_label:.1f} | Elapsed: {elapsed_label}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def main() -> None:
    st.set_page_config(page_title="ZeroNative AI for RTX 5090", layout="wide")

    if not render_access_gate():
        st.stop()

    st.title("ZeroNative AI for RTX 5090")
    st.caption("Ollama + Streamlit with text, Excel, and image inputs")
    WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)

    host, model = render_sidebar()
    if st.sidebar.button("Logout", use_container_width=True):
        st.session_state.app_authenticated = False
        st.rerun()
    temperature = get_model_temperature(model)

    if "history" not in st.session_state:
        st.session_state.history = []
    if "query_cancel_requested" not in st.session_state:
        st.session_state.query_cancel_requested = False
    if "last_elapsed_seconds" not in st.session_state:
        st.session_state.last_elapsed_seconds = None
    if "prepared_downloads" not in st.session_state:
        st.session_state.prepared_downloads = []
    main_col, right_col = st.columns([2.2, 1.0], gap="large")

    with main_col:
        prompt = st.text_area(
            "Prompt",
            height=180,
            placeholder="질문을 입력하세요. 엑셀 파일이나 이미지를 함께 올리면 같이 분석합니다.",
            key="prompt_input",
        )
        query_disabled = not prompt.strip()
        action_col, stop_col, elapsed_col = st.columns([1.2, 1.0, 1.8])
        with action_col:
            query_submitted = st.button("Query", type="primary", disabled=query_disabled, use_container_width=True)
        with stop_col:
            stop_requested = st.button("Stop", disabled=query_disabled, use_container_width=True)
        with elapsed_col:
            elapsed_label = "Elapsed time: -"
            if st.session_state.last_elapsed_seconds is not None:
                elapsed_label = f"Elapsed time: {st.session_state.last_elapsed_seconds:.2f} seconds"
            st.markdown(f"**{elapsed_label}**")

        if stop_requested:
            st.session_state.query_cancel_requested = True
            st.warning("Stop requested. The current query will stop when the next response chunk arrives.")

        uploaded_excels = st.file_uploader(
            "Excel files",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
        )
        uploaded_images = st.file_uploader(
            "Image files",
            type=["png", "jpg", "jpeg", "webp", "bmp"],
            accept_multiple_files=True,
        )

        render_prepared_downloads()

        excel_contexts: list[str] = []
        if uploaded_excels:
            st.subheader("Excel Preview")
            for uploaded_excel in uploaded_excels:
                try:
                    saved_path = save_uploaded_excel(uploaded_excel)
                    context = excel_to_context(uploaded_excel)
                    excel_contexts.append(context)
                    uploaded_excel.seek(0)
                    workbook = pd.ExcelFile(uploaded_excel)
                    st.write(f"Workbook: `{uploaded_excel.name}`")
                    st.caption(f"Saved to: `{saved_path}`")
                    for sheet_name in workbook.sheet_names:
                        frame = workbook.parse(sheet_name)
                        st.write(f"Sheet: `{sheet_name}`")
                        st.dataframe(frame.head(MAX_PREVIEW_ROWS), use_container_width=True)
                except Exception as exc:
                    st.error(f"Failed to read Excel file {uploaded_excel.name}: {exc}")

        image_payloads: list[str] = []
        if uploaded_images:
            st.subheader("Image Preview")
            columns = st.columns(min(len(uploaded_images), 3))
            for index, uploaded_image in enumerate(uploaded_images):
                try:
                    image_base64, image = image_to_base64(uploaded_image)
                    image_payloads.append(image_base64)
                    with columns[index % len(columns)]:
                        st.image(image, caption=uploaded_image.name, use_container_width=True)
                except Exception as exc:
                    st.error(f"Failed to read image {uploaded_image.name}: {exc}")

    with right_col:
        rag_context = render_webdav_rag_panel()

    with main_col:
        if query_submitted:
            with st.spinner("Generating a response..."):
                try:
                    st.session_state.query_cancel_requested = False
                    start_time = time.perf_counter()
                    queried_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    if prompt_requests_tool_list(prompt):
                        answer = build_available_tool_response(model)
                    else:
                        workspace_context = ""
                        workspace_status = None
                        if (not model_supports_tools(model)) and prompt_requests_excel_statistics(prompt):
                            workspace_status = st.info("Excel 통계 계산 중...")
                            workspace_context, excel_status, excel_requested = build_excel_statistics_context_for_prompt(prompt)
                            if excel_requested:
                                if workspace_context.startswith("[Excel Statistics]"):
                                    workspace_status.success(excel_status)
                                else:
                                    workspace_status.warning(excel_status)
                        elif prompt_requests_workspace_scan(prompt):
                            workspace_status = st.info("workspace 스캔 중...")
                            task_file_context, task_file_status, task_file_requested = build_task_file_context(prompt)
                            if task_file_requested:
                                workspace_context = task_file_context
                                if task_file_context.startswith("[Task File]"):
                                    workspace_status.success(task_file_status)
                                else:
                                    workspace_status.warning(task_file_status)
                            else:
                                workspace_context, matched_files = build_workspace_context_for_prompt(prompt)
                                if workspace_context:
                                    workspace_status.success(
                                        f"workspace 스캔 완료: 관련 파일 {matched_files}개를 모델에 전달했습니다."
                                    )
                                else:
                                    workspace_status.warning("workspace에서 관련 파일을 찾지 못했습니다.")
                        answer = call_ollama(
                            host,
                            model,
                            prompt,
                            excel_contexts,
                            rag_context,
                            workspace_context,
                            image_payloads,
                            temperature,
                        )
                    elapsed_seconds = time.perf_counter() - start_time
                    st.session_state.last_elapsed_seconds = elapsed_seconds
                    st.session_state.history.append(
                        {
                            "prompt": prompt,
                            "answer": answer,
                            "elapsed_seconds": elapsed_seconds,
                            "queried_at": queried_at,
                            "model": model,
                            "temperature": temperature,
                        }
                    )
                    st.success("Response received")
                    st.rerun()
                except RuntimeError as exc:
                    elapsed_seconds = time.perf_counter() - start_time
                    st.session_state.last_elapsed_seconds = elapsed_seconds
                    st.warning(str(exc))
                except requests.HTTPError as exc:
                    detail = exc.response.text if exc.response is not None else str(exc)
                    st.error(f"Ollama request failed: {detail}")
                except requests.RequestException as exc:
                    st.error(f"Failed to connect to Ollama at {host}: {exc}")
                except Exception as exc:
                    st.error(f"Unexpected error: {exc}")
                finally:
                    st.session_state.query_cancel_requested = False

        render_workspace_downloads()

        if st.session_state.history:
            st.subheader("History")
            for item in reversed(st.session_state.history):
                render_history_item(item, model, temperature)


if __name__ == "__main__":
    main()
