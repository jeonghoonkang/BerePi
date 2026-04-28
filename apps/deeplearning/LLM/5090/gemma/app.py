from __future__ import annotations

import base64
import io
import json
import os
from pathlib import Path
import shutil
import subprocess
import time
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import pandas as pd
import requests
import streamlit as st
from PIL import Image

DEFAULT_OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://127.0.0.1:11434")
DEFAULT_MODEL = os.getenv("OLLAMA_MODEL", "gemma3:4b")
REQUEST_TIMEOUT = int(os.getenv("OLLAMA_TIMEOUT", "600"))
MAX_PREVIEW_ROWS = 20
APP_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = APP_DIR / "workspace"
MAX_TOOL_FILE_BYTES = 1_000_000
MAX_TOOL_ROUNDS = 8
SUPPORTED_MODEL_OPTIONS = [
    "gemma3:1b",
    "gemma3:4b",
    "gemma3:12b",
    "gemma3:27b",
    "qwen2.5-coder:7b",
]

MODEL_MEMORY_GUIDE_GB = {
    "gemma3:1b": 4,
    "gemma3:4b": 8,
    "gemma3:12b": 20,
    "gemma3:27b": 40,
    "qwen2.5-coder:7b": 8,
}

FILE_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": "List files and directories inside the workspace directory.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Relative path inside the workspace", "default": "."},
                    "recursive": {"type": "boolean", "description": "Whether to include nested files", "default": False},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "Read a UTF-8 text file from the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["path"],
                "properties": {
                    "path": {"type": "string", "description": "Relative file path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": "Write UTF-8 text content to a file inside the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["path", "content"],
                "properties": {
                    "path": {"type": "string", "description": "Relative file path inside the workspace"},
                    "content": {"type": "string", "description": "Text content to write to the file"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "copy_path",
            "description": "Copy a file or directory inside the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["source_path", "destination_path"],
                "properties": {
                    "source_path": {"type": "string", "description": "Relative source path inside the workspace"},
                    "destination_path": {"type": "string", "description": "Relative destination path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_path",
            "description": "Delete a file or directory inside the workspace directory.",
            "parameters": {
                "type": "object",
                "required": ["path"],
                "properties": {
                    "path": {"type": "string", "description": "Relative path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_workbook_info",
            "description": "Get workbook sheet names and basic dimensions for an Excel file in the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_sheet_preview",
            "description": "Preview the first rows of a sheet in an Excel file from the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to preview"},
                    "max_rows": {"type": "integer", "description": "Number of rows to preview", "default": 20},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_read_cells",
            "description": "Read a cell or cell range from an Excel sheet in the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name", "cell_range"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to read"},
                    "cell_range": {"type": "string", "description": "Excel range like A1 or A1:C10"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_write_cell",
            "description": "Write a value into a single Excel cell in the workspace.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name", "cell", "value"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to update"},
                    "cell": {"type": "string", "description": "Target cell like B3"},
                    "value": {"type": "string", "description": "Value to write"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_aggregate_range",
            "description": "Run numeric aggregation on an Excel range. Supported operations: sum, average, min, max, count.",
            "parameters": {
                "type": "object",
                "required": ["path", "sheet_name", "cell_range", "operation"],
                "properties": {
                    "path": {"type": "string", "description": "Relative Excel file path inside the workspace"},
                    "sheet_name": {"type": "string", "description": "Sheet name to aggregate"},
                    "cell_range": {"type": "string", "description": "Excel range like A1:C10"},
                    "operation": {"type": "string", "description": "sum, average, min, max, or count"},
                },
            },
        },
    },
]


def excel_to_context(uploaded_file) -> str:
    """Create a concise, prompt-friendly summary from an Excel workbook."""
    uploaded_file.seek(0)
    workbook = pd.ExcelFile(uploaded_file)
    sections: list[str] = []

    for sheet_name in workbook.sheet_names:
        frame = workbook.parse(sheet_name)
        preview = frame.head(MAX_PREVIEW_ROWS).fillna("")
        preview_csv = preview.to_csv(index=False)
        sections.append(
            "\n".join(
                [
                    f"[Sheet] {sheet_name}",
                    f"Rows: {len(frame)}",
                    f"Columns: {len(frame.columns)}",
                    "Column names: " + ", ".join(str(column) for column in frame.columns),
                    f"Preview ({min(len(preview), MAX_PREVIEW_ROWS)} rows):",
                    preview_csv.strip(),
                ]
            )
        )

    return "\n\n".join(sections)


def image_to_base64(uploaded_file) -> tuple[str, Image.Image]:
    """Return a base64 string for Ollama and a PIL preview image."""
    data = uploaded_file.getvalue()
    encoded = base64.b64encode(data).decode("utf-8")
    image = Image.open(io.BytesIO(data))
    return encoded, image


def save_uploaded_excel(uploaded_file) -> Path:
    """Persist an uploaded Excel file to the workspace directory."""
    WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)
    destination = WORKSPACE_DIR / Path(uploaded_file.name).name
    destination.write_bytes(uploaded_file.getvalue())
    return destination


def safe_excel_path(relative_path: str) -> Path:
    """Resolve a workspace path and require an Excel file extension."""
    path = safe_workspace_path(relative_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
        raise ValueError(f"Not an Excel file: {relative_path}")
    return path


def safe_workspace_path(relative_path: str) -> Path:
    """Resolve a user path while preventing access outside the workspace."""
    raw_path = relative_path.strip() or "."
    candidate = (WORKSPACE_DIR / raw_path).resolve()
    workspace_root = WORKSPACE_DIR.resolve()
    if candidate != workspace_root and workspace_root not in candidate.parents:
        raise ValueError("Access outside the workspace directory is not allowed.")
    return candidate


def workspace_relative(path: Path) -> str:
    """Return a workspace-relative display path."""
    return str(path.resolve().relative_to(WORKSPACE_DIR.resolve()))


def list_workspace_entries(relative_path: str = ".", recursive: bool = False) -> str:
    """List files and directories within the workspace."""
    path = safe_workspace_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"Path not found: {relative_path}")
    if not path.is_dir():
        raise NotADirectoryError(f"Not a directory: {relative_path}")

    if recursive:
        entries = sorted(path.rglob("*"))
    else:
        entries = sorted(path.iterdir())

    lines = []
    for entry in entries[:200]:
        entry_type = "dir" if entry.is_dir() else "file"
        lines.append(f"{entry_type}: {workspace_relative(entry)}")
    if len(entries) > 200:
        lines.append(f"... truncated, total entries: {len(entries)}")
    return "\n".join(lines) if lines else "(empty directory)"


def read_workspace_file(relative_path: str) -> str:
    """Read a text file from the workspace."""
    path = safe_workspace_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {relative_path}")
    if not path.is_file():
        raise IsADirectoryError(f"Path is not a file: {relative_path}")
    if path.stat().st_size > MAX_TOOL_FILE_BYTES:
        raise ValueError(f"File too large to read safely: {relative_path}")
    return path.read_text(encoding="utf-8", errors="replace")


def write_workspace_file(relative_path: str, content: str) -> str:
    """Write text content to a workspace file."""
    path = safe_workspace_path(relative_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return f"Saved file: {workspace_relative(path)}"


def copy_workspace_path(source_path: str, destination_path: str) -> str:
    """Copy a workspace file or directory."""
    source = safe_workspace_path(source_path)
    destination = safe_workspace_path(destination_path)
    if not source.exists():
        raise FileNotFoundError(f"Source not found: {source_path}")
    destination.parent.mkdir(parents=True, exist_ok=True)

    if source.is_dir():
        shutil.copytree(source, destination, dirs_exist_ok=True)
        return f"Copied directory: {workspace_relative(source)} -> {workspace_relative(destination)}"

    shutil.copy2(source, destination)
    return f"Copied file: {workspace_relative(source)} -> {workspace_relative(destination)}"


def delete_workspace_path(relative_path: str) -> str:
    """Delete a workspace file or directory."""
    path = safe_workspace_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"Path not found: {relative_path}")
    if path == WORKSPACE_DIR.resolve():
        raise ValueError("Deleting the workspace root is not allowed.")

    if path.is_dir():
        shutil.rmtree(path)
        return f"Deleted directory: {relative_path}"

    path.unlink()
    return f"Deleted file: {relative_path}"


def excel_workbook_info(relative_path: str) -> str:
    """Return workbook sheet metadata."""
    path = safe_excel_path(relative_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {relative_path}")

    workbook = load_workbook(path, data_only=False)
    lines = [f"Workbook: {workspace_relative(path)}"]
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines.append(
            f"Sheet: {sheet_name} | rows={sheet.max_row} | cols={sheet.max_column}"
        )
    workbook.close()
    return "\n".join(lines)


def excel_sheet_preview(relative_path: str, sheet_name: str, max_rows: int = 20) -> str:
    """Preview top rows from an Excel sheet."""
    path = safe_excel_path(relative_path)
    preview_rows = max(1, min(int(max_rows), 100))
    frame = pd.read_excel(path, sheet_name=sheet_name).head(preview_rows).fillna("")
    return frame.to_csv(index=False).strip()


def excel_read_cells(relative_path: str, sheet_name: str, cell_range: str) -> str:
    """Read one or more cells from an Excel sheet."""
    path = safe_excel_path(relative_path)
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name]

    if ":" not in cell_range:
        value = sheet[cell_range].value
        workbook.close()
        return f"{cell_range} = {value}"

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    lines = []
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=False,
    ):
        row_items = [f"{cell.coordinate}={cell.value}" for cell in row]
        lines.append(", ".join(row_items))
    workbook.close()
    return "\n".join(lines)


def coerce_excel_value(value: str):
    """Convert text input into a simple Excel cell value."""
    lowered = value.strip().lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return value


def excel_write_cell(relative_path: str, sheet_name: str, cell: str, value: str) -> str:
    """Write a value into a single Excel cell."""
    path = safe_excel_path(relative_path)
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    sheet[cell] = coerce_excel_value(value)
    workbook.save(path)
    workbook.close()
    return f"Updated {workspace_relative(path)} {sheet_name}!{cell} = {value}"


def excel_aggregate_range(relative_path: str, sheet_name: str, cell_range: str, operation: str) -> str:
    """Aggregate numeric values from an Excel range."""
    path = safe_excel_path(relative_path)
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    numbers = []
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        for value in row:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numbers.append(float(value))
    workbook.close()

    if operation not in {"sum", "average", "min", "max", "count"}:
        raise ValueError(f"Unsupported operation: {operation}")
    if operation != "count" and not numbers:
        raise ValueError(f"No numeric values found in range: {cell_range}")

    if operation == "sum":
        result = sum(numbers)
    elif operation == "average":
        result = sum(numbers) / len(numbers)
    elif operation == "min":
        result = min(numbers)
    elif operation == "max":
        result = max(numbers)
    else:
        result = len(numbers)

    return f"{operation}({sheet_name}!{cell_range}) = {result}"


def execute_file_tool(name: str, arguments: dict) -> str:
    """Run a validated workspace tool call and return its result."""
    if name == "list_files":
        return list_workspace_entries(
            relative_path=str(arguments.get("path", ".")),
            recursive=bool(arguments.get("recursive", False)),
        )
    if name == "read_file":
        return read_workspace_file(str(arguments["path"]))
    if name == "write_file":
        return write_workspace_file(str(arguments["path"]), str(arguments["content"]))
    if name == "copy_path":
        return copy_workspace_path(
            source_path=str(arguments["source_path"]),
            destination_path=str(arguments["destination_path"]),
        )
    if name == "delete_path":
        return delete_workspace_path(str(arguments["path"]))
    if name == "excel_workbook_info":
        return excel_workbook_info(str(arguments["path"]))
    if name == "excel_sheet_preview":
        return excel_sheet_preview(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            max_rows=int(arguments.get("max_rows", 20)),
        )
    if name == "excel_read_cells":
        return excel_read_cells(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            cell_range=str(arguments["cell_range"]),
        )
    if name == "excel_write_cell":
        return excel_write_cell(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            cell=str(arguments["cell"]),
            value=str(arguments["value"]),
        )
    if name == "excel_aggregate_range":
        return excel_aggregate_range(
            relative_path=str(arguments["path"]),
            sheet_name=str(arguments["sheet_name"]),
            cell_range=str(arguments["cell_range"]),
            operation=str(arguments["operation"]).lower(),
        )
    raise ValueError(f"Unknown tool: {name}")


def normalize_tool_arguments(arguments) -> dict:
    """Normalize tool arguments that may arrive as a dict or JSON string."""
    if isinstance(arguments, dict):
        return arguments
    if isinstance(arguments, str):
        parsed = json.loads(arguments)
        if isinstance(parsed, dict):
            return parsed
    raise ValueError(f"Unsupported tool arguments: {arguments}")


def build_user_message(prompt: str, excel_contexts: Iterable[str]) -> str:
    """Build the final prompt content sent to the model."""
    content_parts = [prompt.strip()]

    excel_contexts = [context for context in excel_contexts if context.strip()]
    if excel_contexts:
        content_parts.append(
            "The user also uploaded Excel data. Use the workbook summaries below when relevant.\n\n"
            + "\n\n".join(excel_contexts)
        )
    content_parts.append(
        f"You may use workspace tools when needed. The workspace root is: {WORKSPACE_DIR.resolve()}"
    )

    return "\n\n".join(part for part in content_parts if part)


def call_ollama(host: str, model: str, prompt: str, excel_contexts: list[str], images: list[str]) -> str:
    """Send a chat request to Ollama and allow validated workspace tool calls."""
    url = f"{host.rstrip('/')}/api/chat"
    messages = [
        {
            "role": "system",
            "content": (
                "You are a helpful assistant running on an RTX 5090 server. "
                "Answer clearly, use uploaded Excel context when provided, analyze images when attached, "
                "and use workspace file tools when they are needed to complete the user's request. "
                "Never claim you changed files unless a tool call succeeded."
            ),
        },
        {
            "role": "user",
            "content": build_user_message(prompt, excel_contexts),
            "images": images,
        },
    ]

    for _ in range(MAX_TOOL_ROUNDS):
        if st.session_state.get("query_cancel_requested"):
            raise RuntimeError("Query stopped by user.")

        response = requests.post(
            url,
            json={
                "model": model,
                "stream": False,
                "messages": messages,
                "tools": FILE_TOOLS,
            },
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        event = response.json()
        message = event.get("message", {})
        messages.append(message)
        tool_calls = message.get("tool_calls") or []

        if not tool_calls:
            return message.get("content", "").strip()

        for tool_call in tool_calls:
            if st.session_state.get("query_cancel_requested"):
                raise RuntimeError("Query stopped by user.")

            function = tool_call.get("function", {})
            tool_name = function.get("name", "")
            arguments = normalize_tool_arguments(function.get("arguments", {}) or {})

            try:
                result = execute_file_tool(tool_name, arguments)
            except Exception as exc:
                result = f"Tool error: {exc}"

            messages.append(
                {
                    "role": "tool",
                    "tool_name": tool_name,
                    "content": str(result),
                }
            )

    raise RuntimeError("Tool calling stopped after reaching the maximum tool round limit.")


def fetch_installed_models(host: str) -> dict[str, dict]:
    """Return installed Ollama models keyed by name."""
    response = requests.get(f"{host.rstrip('/')}/api/tags", timeout=30)
    response.raise_for_status()
    data = response.json()
    models = data.get("models", [])
    return {
        model.get("name", "").strip(): model
        for model in models
        if model.get("name", "").strip()
    }


def fetch_model_show_info(host: str, model: str) -> dict:
    """Return detailed model information from Ollama show API."""
    response = requests.post(
        f"{host.rstrip('/')}/api/show",
        json={"model": model},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def detect_gpu_info() -> tuple[str | None, int | None, str | None]:
    """Detect GPU name and memory in GiB using nvidia-smi when available."""
    if not shutil.which("nvidia-smi"):
        return None, None, "nvidia-smi not found"

    try:
        result = subprocess.run(
            [
                "nvidia-smi",
                "--query-gpu=name,memory.total",
                "--format=csv,noheader,nounits",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
    except (subprocess.CalledProcessError, OSError) as exc:
        return None, None, str(exc)

    first_line = result.stdout.strip().splitlines()[0] if result.stdout.strip() else ""
    if not first_line:
        return None, None, "No GPU information returned"

    parts = [part.strip() for part in first_line.split(",", maxsplit=1)]
    if len(parts) != 2:
        return None, None, f"Unexpected nvidia-smi output: {first_line}"

    gpu_name = parts[0]
    try:
        memory_mib = int(parts[1])
    except ValueError:
        return gpu_name, None, f"Unexpected memory value: {parts[1]}"

    memory_gib = max(memory_mib // 1024, 1)
    return gpu_name, memory_gib, None


def recommend_models_for_gpu(memory_gib: int | None) -> tuple[str | None, list[str]]:
    """Return a recommended default model and compatible model list."""
    if memory_gib is None:
        return None, []

    compatible_models = [
        model_name
        for model_name, required_gb in MODEL_MEMORY_GUIDE_GB.items()
        if memory_gib >= required_gb
    ]

    if not compatible_models:
        return "gemma3:1b", ["gemma3:1b"]

    return compatible_models[-1], compatible_models


def format_bytes(size_bytes: int | None) -> str:
    """Format a byte count into a readable string."""
    if size_bytes is None:
        return "Unknown"

    value = float(size_bytes)
    units = ["B", "KB", "MB", "GB", "TB"]
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.2f} {unit}"
        value /= 1024
    return f"{size_bytes} B"


def format_duration(seconds: float | None) -> str:
    """Format seconds into a short readable duration."""
    if seconds is None or seconds < 0:
        return "Unknown"

    total_seconds = int(seconds)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)

    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def get_default_ollama_models_root() -> Path:
    """Return the configured or default local Ollama model root."""
    configured = os.getenv("OLLAMA_MODELS")
    if configured:
        return Path(configured).expanduser()
    return Path.home() / ".ollama" / "models"


def get_selected_ollama_models_root() -> Path:
    """Return the user-selected model root or the default root."""
    selected = st.session_state.get("ollama_models_path")
    if selected:
        return Path(selected).expanduser()
    return get_default_ollama_models_root()


def move_model_storage(source_root: Path, destination_root: Path) -> str:
    """Move existing Ollama model files to a new storage root."""
    source = source_root.expanduser().resolve()
    destination = destination_root.expanduser().resolve()

    if source == destination:
        return f"Model storage path is already set to: {destination}"

    if not source.exists():
        destination.mkdir(parents=True, exist_ok=True)
        return f"Source model path does not exist yet. Created destination: {destination}"

    destination.mkdir(parents=True, exist_ok=True)
    moved_items = []
    for child in source.iterdir():
        target = destination / child.name
        shutil.move(str(child), str(target))
        moved_items.append(child.name)

    return (
        f"Moved {len(moved_items)} item(s) from {source} to {destination}. "
        "Restart Ollama with OLLAMA_MODELS pointing to the new path to use it for future downloads."
    )


def get_model_storage_path(model: str) -> Path:
    """Infer the local Ollama manifest path for a model tag."""
    registry = "registry.ollama.ai"
    namespace = "library"
    repository = model
    tag = "latest"

    if ":" in model:
        repository, tag = model.rsplit(":", maxsplit=1)

    if "/" in repository:
        namespace, repository = repository.split("/", maxsplit=1)

    return get_selected_ollama_models_root() / "manifests" / registry / namespace / repository / tag


def get_selected_model_info(host: str, model: str, installed_model_map: dict[str, dict]) -> tuple[dict | None, str | None]:
    """Return selected model metadata from tags and show API."""
    installed_info = installed_model_map.get(model)
    try:
        show_info = fetch_model_show_info(host, model)
    except requests.RequestException as exc:
        if installed_info:
            return {
                "installed": installed_info,
                "show": {},
            }, str(exc)
        return None, str(exc)
    return {
        "installed": installed_info,
        "show": show_info,
    }, None


def pull_model(host: str, model: str, status_placeholder, progress_placeholder) -> None:
    """Download a model from Ollama and show progress in the sidebar."""
    response = requests.post(
        f"{host.rstrip('/')}/api/pull",
        json={"model": model, "stream": True},
        timeout=REQUEST_TIMEOUT,
        stream=True,
    )
    response.raise_for_status()

    progress_bar = progress_placeholder.progress(0)
    last_status = "Starting download..."
    status_placeholder.info(last_status)
    started_at = time.perf_counter()

    for raw_line in response.iter_lines():
        if not raw_line:
            continue

        event = json.loads(raw_line.decode("utf-8"))
        status_text = event.get("status", last_status)
        total = event.get("total")
        completed = event.get("completed")

        if total and completed is not None and total > 0:
            ratio = min(max(completed / total, 0.0), 1.0)
            elapsed_seconds = max(time.perf_counter() - started_at, 0.001)
            bytes_per_second = completed / elapsed_seconds if completed > 0 else 0.0
            remaining_bytes = max(total - completed, 0)
            eta_seconds = remaining_bytes / bytes_per_second if bytes_per_second > 0 else None
            progress_text = (
                f"{status_text} | {ratio * 100:.1f}% | "
                f"{format_bytes(completed)} / {format_bytes(total)} | "
                f"ETA {format_duration(eta_seconds)}"
            )
            progress_bar.progress(ratio, text=progress_text)
        else:
            progress_bar.progress(0, text=status_text)

        last_status = status_text
        status_placeholder.info(status_text)

    progress_bar.progress(1.0, text="Download complete")
    status_placeholder.success(f"Model download completed: {model}")


def render_sidebar() -> tuple[str, str]:
    st.sidebar.header("Runtime")
    host = st.sidebar.text_input("Ollama Host", value=DEFAULT_OLLAMA_HOST)
    st.sidebar.write(f"Workspace: `{WORKSPACE_DIR.resolve()}`")

    default_model_root = get_default_ollama_models_root()
    if "ollama_models_path" not in st.session_state:
        st.session_state.ollama_models_path = str(default_model_root)
    if "model_path_message" not in st.session_state:
        st.session_state.model_path_message = ""

    model_root_input = st.sidebar.text_input(
        "Model Download Path",
        value=st.session_state.ollama_models_path,
        help="This path is used for display and migration. Ollama must be restarted with OLLAMA_MODELS set to this path for future downloads to use it.",
    )
    apply_model_root = st.sidebar.button("Use Model Path", use_container_width=True)
    move_model_root = st.sidebar.button("Move Existing Model Files", use_container_width=True)

    if apply_model_root:
        st.session_state.ollama_models_path = model_root_input.strip() or str(default_model_root)
        st.session_state.model_path_message = (
            "Model path updated in the app. Restart Ollama with OLLAMA_MODELS set to this path for future downloads."
        )
        st.rerun()

    if move_model_root:
        try:
            source_root = get_selected_ollama_models_root()
            destination_root = Path(model_root_input.strip() or str(default_model_root))
            message = move_model_storage(source_root, destination_root)
            st.session_state.ollama_models_path = str(destination_root.expanduser())
            st.session_state.model_path_message = message
            st.rerun()
        except Exception as exc:
            st.session_state.model_path_message = f"Model file move failed: {exc}"

    if st.session_state.model_path_message:
        st.sidebar.info(st.session_state.model_path_message)

    refresh_models = st.sidebar.button("Refresh Installed Models", use_container_width=True)

    if refresh_models or "installed_models" not in st.session_state:
        try:
            st.session_state.installed_models = fetch_installed_models(host)
            st.session_state.model_error = ""
        except requests.RequestException as exc:
            st.session_state.installed_models = {}
            st.session_state.model_error = f"Failed to load installed models: {exc}"

    installed_model_map = st.session_state.get("installed_models", {})
    installed_models = sorted(installed_model_map.keys())
    model_candidates = list(dict.fromkeys(SUPPORTED_MODEL_OPTIONS + installed_models))
    gpu_name, gpu_memory_gib, gpu_error = detect_gpu_info()
    recommended_model, compatible_models = recommend_models_for_gpu(gpu_memory_gib)

    if "preferred_model" not in st.session_state:
        st.session_state.preferred_model = recommended_model or DEFAULT_MODEL

    if "auto_select_downloaded_model" not in st.session_state:
        st.session_state.auto_select_downloaded_model = True

    if st.session_state.get("selected_model") not in model_candidates:
        preferred = st.session_state.get("preferred_model", DEFAULT_MODEL)
        if preferred in model_candidates:
            st.session_state.selected_model = preferred
        elif DEFAULT_MODEL in model_candidates:
            st.session_state.selected_model = DEFAULT_MODEL
        else:
            st.session_state.selected_model = model_candidates[0]

    st.sidebar.subheader("GPU Recommendation")
    if gpu_name and gpu_memory_gib:
        st.sidebar.write(f"GPU: `{gpu_name}`")
        st.sidebar.write(f"Memory: `{gpu_memory_gib} GiB`")
        if recommended_model:
            st.sidebar.success(f"Recommended default: `{recommended_model}`")
        if compatible_models:
            st.sidebar.caption("Fits this GPU")
            for compatible_model in compatible_models:
                st.sidebar.code(compatible_model)
    else:
        st.sidebar.info("GPU memory could not be detected automatically.")
        if gpu_error:
            st.sidebar.caption(gpu_error)

    selected_model = st.sidebar.selectbox(
        "Model Select",
        options=model_candidates,
        index=model_candidates.index(st.session_state.selected_model),
    )
    st.session_state.selected_model = selected_model

    custom_model = st.sidebar.text_input("Custom Model Tag", value=selected_model)
    model = custom_model.strip() or selected_model

    if st.session_state.get("model_error"):
        st.sidebar.warning(st.session_state.model_error)

    with st.sidebar.expander("Installed Models", expanded=False):
        if installed_models:
            for installed_model in installed_models:
                st.write(f"- {installed_model}")
        else:
            st.write("No installed models found.")

    st.session_state.auto_select_downloaded_model = st.sidebar.checkbox(
        "Auto-select downloaded model",
        value=st.session_state.auto_select_downloaded_model,
    )

    if recommended_model and st.sidebar.button("Use Recommended Model", use_container_width=True):
        st.session_state.preferred_model = recommended_model
        st.session_state.selected_model = recommended_model
        st.rerun()

    st.sidebar.markdown("Recommended model options")
    for option in SUPPORTED_MODEL_OPTIONS:
        required_gb = MODEL_MEMORY_GUIDE_GB.get(option)
        label = f"{option}  ({required_gb} GiB+ recommended)" if required_gb else option
        st.sidebar.code(label)

    st.sidebar.subheader("Current Model")
    selected_model_info, selected_model_error = get_selected_model_info(host, model, installed_model_map)
    storage_root = get_selected_ollama_models_root()
    storage_path = get_model_storage_path(model)
    st.sidebar.write(f"Model: `{model}`")
    st.sidebar.write(f"Storage root: `{storage_root}`")
    st.sidebar.write(f"Storage path: `{storage_path}`")
    st.sidebar.caption("Future downloads use this path only after Ollama is restarted with OLLAMA_MODELS set to the same directory.")

    if selected_model_info:
        installed_info = selected_model_info.get("installed") or {}
        show_info = selected_model_info.get("show") or {}
        details = installed_info.get("details") or {}
        model_size = installed_info.get("size")
        parameter_size = details.get("parameter_size") or show_info.get("details", {}).get("parameter_size")
        quantization = details.get("quantization_level") or show_info.get("details", {}).get("quantization_level")
        family = details.get("family") or show_info.get("details", {}).get("family")

        st.sidebar.write(f"Model size: `{format_bytes(model_size)}`")
        if parameter_size:
            st.sidebar.write(f"Parameters: `{parameter_size}`")
        if quantization:
            st.sidebar.write(f"Quantization: `{quantization}`")
        if family:
            st.sidebar.write(f"Family: `{family}`")
    elif selected_model_error:
        st.sidebar.caption(f"Model detail lookup failed: {selected_model_error}")

    status_placeholder = st.sidebar.empty()
    progress_placeholder = st.sidebar.empty()
    if st.sidebar.button("Download Selected Model", type="primary", use_container_width=True):
        try:
            pull_model(host, model, status_placeholder, progress_placeholder)
            st.session_state.installed_models = fetch_installed_models(host)
            st.session_state.preferred_model = model
            if st.session_state.auto_select_downloaded_model:
                st.session_state.selected_model = model
            st.rerun()
        except requests.RequestException as exc:
            status_placeholder.error(f"Model download failed: {exc}")

    st.sidebar.markdown("Example: `OLLAMA_HOST=http://127.0.0.1:11434 ollama serve`")

    return host, model


def main() -> None:
    st.set_page_config(page_title="Gemma 3 4B on RTX 5090", layout="wide")
    st.title("Gemma 3 4B AI for RTX 5090")
    st.caption("Ollama + Streamlit with text, Excel, and image inputs")
    WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)

    host, model = render_sidebar()

    if "history" not in st.session_state:
        st.session_state.history = []
    if "query_cancel_requested" not in st.session_state:
        st.session_state.query_cancel_requested = False
    if "last_elapsed_seconds" not in st.session_state:
        st.session_state.last_elapsed_seconds = None

    prompt = st.text_area(
        "Prompt",
        height=180,
        placeholder="질문을 입력하세요. 엑셀 파일이나 이미지를 함께 올리면 같이 분석합니다.",
    )
    query_disabled = not prompt.strip()
    action_col, stop_col, elapsed_col = st.columns([1.2, 1.0, 1.8])
    with action_col:
        query_submitted = st.button("Query Gemma", type="primary", disabled=query_disabled, use_container_width=True)
    with stop_col:
        stop_requested = st.button("Stop", disabled=query_disabled, use_container_width=True)
    with elapsed_col:
        elapsed_label = "Elapsed time: -"
        if st.session_state.last_elapsed_seconds is not None:
            elapsed_label = f"Elapsed time: {st.session_state.last_elapsed_seconds:.2f} seconds"
        st.markdown(f"**{elapsed_label}**")

    if stop_requested:
        st.session_state.query_cancel_requested = True
        st.warning("Stop requested. The current query will stop when the next response chunk arrives.")

    uploaded_excels = st.file_uploader(
        "Excel files",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )
    uploaded_images = st.file_uploader(
        "Image files",
        type=["png", "jpg", "jpeg", "webp", "bmp"],
        accept_multiple_files=True,
    )

    excel_contexts: list[str] = []
    if uploaded_excels:
        st.subheader("Excel Preview")
        for uploaded_excel in uploaded_excels:
            try:
                saved_path = save_uploaded_excel(uploaded_excel)
                context = excel_to_context(uploaded_excel)
                excel_contexts.append(context)
                uploaded_excel.seek(0)
                workbook = pd.ExcelFile(uploaded_excel)
                st.write(f"Workbook: `{uploaded_excel.name}`")
                st.caption(f"Saved to: `{saved_path}`")
                for sheet_name in workbook.sheet_names:
                    frame = workbook.parse(sheet_name)
                    st.write(f"Sheet: `{sheet_name}`")
                    st.dataframe(frame.head(MAX_PREVIEW_ROWS), use_container_width=True)
            except Exception as exc:
                st.error(f"Failed to read Excel file {uploaded_excel.name}: {exc}")

    image_payloads: list[str] = []
    if uploaded_images:
        st.subheader("Image Preview")
        columns = st.columns(min(len(uploaded_images), 3))
        for index, uploaded_image in enumerate(uploaded_images):
            try:
                image_base64, image = image_to_base64(uploaded_image)
                image_payloads.append(image_base64)
                with columns[index % len(columns)]:
                    st.image(image, caption=uploaded_image.name, use_container_width=True)
            except Exception as exc:
                st.error(f"Failed to read image {uploaded_image.name}: {exc}")

    if query_submitted:
        with st.spinner("Gemma is generating a response..."):
            try:
                st.session_state.query_cancel_requested = False
                start_time = time.perf_counter()
                answer = call_ollama(host, model, prompt, excel_contexts, image_payloads)
                elapsed_seconds = time.perf_counter() - start_time
                st.session_state.last_elapsed_seconds = elapsed_seconds
                st.session_state.history.append(
                    {
                        "prompt": prompt,
                        "answer": answer,
                        "elapsed_seconds": elapsed_seconds,
                        "model": model,
                    }
                )
                st.success("Response received")
                st.rerun()
            except RuntimeError as exc:
                elapsed_seconds = time.perf_counter() - start_time
                st.session_state.last_elapsed_seconds = elapsed_seconds
                st.warning(str(exc))
            except requests.HTTPError as exc:
                detail = exc.response.text if exc.response is not None else str(exc)
                st.error(f"Ollama request failed: {detail}")
            except requests.RequestException as exc:
                st.error(f"Failed to connect to Ollama at {host}: {exc}")
            except Exception as exc:
                st.error(f"Unexpected error: {exc}")
            finally:
                st.session_state.query_cancel_requested = False

    if st.session_state.history:
        st.subheader("History")
        for item in reversed(st.session_state.history):
            st.markdown("**Prompt**")
            st.write(item["prompt"])
            st.markdown("**Answer**")
            st.write(item["answer"])
            if item.get("elapsed_seconds") is not None:
                st.caption(
                    f"Model: {item.get('model', model)} | Elapsed: {item['elapsed_seconds']:.2f} seconds"
                )


if __name__ == "__main__":
    main()
